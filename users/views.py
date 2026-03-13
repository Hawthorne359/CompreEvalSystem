"""
认证与用户 API。
"""
import uuid
import json

from django.core.cache import cache
from django.db.models import Max, Q
from django.utils import timezone
from rest_framework import status
from rest_framework.permissions import IsAuthenticated, AllowAny
from rest_framework.response import Response
from rest_framework.views import APIView
from rest_framework.viewsets import ModelViewSet, ReadOnlyModelViewSet
from rest_framework.decorators import action
from rest_framework_simplejwt.views import TokenObtainPairView, TokenRefreshView
from rest_framework_simplejwt.serializers import TokenObtainPairSerializer
from rest_framework.parsers import MultiPartParser, FormParser

from audit.models import OperationLog
from audit.services import log_action
from .models import User, Role, UserRole, ImportedUserBatch
from .role_resolver import (
    ROLE_LEVEL_ASSISTANT, ROLE_LEVEL_COUNSELOR, ROLE_LEVEL_DIRECTOR, ROLE_LEVEL_SUPERADMIN,
    get_role_display_name,
)
from .serializers import (
    UserSerializer,
    UserListSerializer,
    UserCreateUpdateSerializer,
    RoleSerializer,
    ImportedUserBatchSerializer,
)

# 密码确认 token 在 Django cache 中的有效期（秒）
_VERIFY_TOKEN_TTL = 300  # 5 分钟
_IMPORT_PREVIEW_TTL = 1800  # 30 分钟


class CustomTokenObtainPairSerializer(TokenObtainPairSerializer):
    """登录成功后在响应中附带用户与角色信息。"""

    def validate(self, attrs):
        data = super().validate(attrs)
        data['user'] = UserSerializer(self.user).data
        return data


class CustomTokenObtainPairView(TokenObtainPairView):
    """登录：返回 access/refresh 及当前用户信息；写 INFO 级操作日志；生成 SSO session_key。"""
    serializer_class = CustomTokenObtainPairSerializer

    def post(self, request, *args, **kwargs):
        response = super().post(request, *args, **kwargs)
        if response.status_code == 200:
            # 登录成功后，通过用户名反查用户对象写日志
            username = request.data.get('username', '')
            user_obj = User.objects.filter(username=username).first()
            log_action(
                user=user_obj,
                action='login',
                module=OperationLog.MODULE_AUTH,
                level=OperationLog.LEVEL_INFO,
                target_type='user',
                target_id=user_obj.id if user_obj else None,
                target_repr=username,
                request=request,
            )
            # 生成 SSO session_key，写入 Cache 以踢出旧会话（同账号新登录自动失效旧端）
            if user_obj:
                session_key = str(uuid.uuid4())
                cache.set(
                    f'sso_session:{user_obj.id}',
                    session_key,
                    timeout=7 * 24 * 3600,  # 与 refresh token 同寿命
                )
                response.data['session_key'] = session_key
                # SSE 实时推送：通知旧设备会话已被替换，携带新登录设备信息
                from realtime.registry import publish
                from realtime.utils import parse_user_agent, geolocate_ip, get_request_ip
                ua = request.META.get('HTTP_USER_AGENT', '')
                ip = get_request_ip(request)
                device = parse_user_agent(ua)
                location = geolocate_ip(ip)
                publish(user_obj.id, {
                    'type': 'session_replaced',
                    'device': device,
                    'location': location,
                    'ip': ip,
                })
        return response


class AuthLoginView(CustomTokenObtainPairView):
    """POST /api/v1/auth/login/ 登录。"""
    permission_classes = [AllowAny]


class AuthRefreshView(TokenRefreshView):
    """POST /api/v1/auth/refresh/ 刷新 token。"""
    permission_classes = [AllowAny]



class AuthLogoutView(APIView):
    """
    POST /api/v1/auth/logout/ 登出留痕。
    JWT 无状态，客户端只需删除本地 token；此端点负责记录日志并清除 SSO session_key。
    """
    permission_classes = [IsAuthenticated]

    def post(self, request):
        # 主动清除 SSO session_key，使该账号的所有端会话立即失效
        cache.delete(f'sso_session:{request.user.id}')
        log_action(
            user=request.user,
            action='logout',
            module=OperationLog.MODULE_AUTH,
            level=OperationLog.LEVEL_INFO,
            target_type='user',
            target_id=request.user.id,
            target_repr=request.user.username,
            request=request,
        )
        return Response({'detail': '已登出'})


class AuthSwitchRoleView(APIView):
    """
    POST /api/v1/auth/switch-role/
    切换当前角色；当目标角色等级高于当前角色等级时，需传 password 二次验证。
    """
    permission_classes = [IsAuthenticated]

    def post(self, request):
        role_id = request.data.get('role_id')
        password = request.data.get('password', '')
        if role_id is None:
            return Response(
                {'detail': '请提供 role_id'},
                status=status.HTTP_400_BAD_REQUEST,
            )
        try:
            role = Role.objects.get(pk=role_id)
        except Role.DoesNotExist:
            return Response(
                {'detail': '角色不存在'},
                status=status.HTTP_404_NOT_FOUND,
            )
        user = request.user
        if not UserRole.objects.filter(user=user, role=role).exists():
            return Response(
                {'detail': '您不拥有该角色'},
                status=status.HTTP_403_FORBIDDEN,
            )
        current_level = user.current_role.level if user.current_role else -1
        target_level = role.level if role else -1
        need_password = target_level > current_level

        if need_password:
            if not password:
                return Response(
                    {'detail': '升权切换需输入密码验证'},
                    status=status.HTTP_400_BAD_REQUEST,
                )
            if not user.check_password(password):
                return Response(
                    {'detail': '密码错误'},
                    status=status.HTTP_401_UNAUTHORIZED,
                )
        old_role_name = user.current_role.name if user.current_role else '无'
        old_level = current_level
        is_downgrade = target_level < current_level
        user.current_role = role
        user.save(update_fields=['current_role'])
        log_action(
            user=user,
            action='switch_role',
            module=OperationLog.MODULE_AUTH,
            level=OperationLog.LEVEL_WARNING if is_downgrade else OperationLog.LEVEL_INFO,
            target_type='role',
            target_id=role.id,
            target_repr=role.name,
            extra={
                'from_role': old_role_name,
                'to_role': role.name,
                'from_level': old_level,
                'to_level': target_level,
                'is_downgrade': is_downgrade,
            },
            request=request,
        )
        return Response({
            'detail': '切换成功',
            'current_role': RoleSerializer(role).data,
            'user': UserSerializer(user).data,
        })


class VerifyPasswordView(APIView):
    """
    POST /api/v1/auth/verify-password/
    验证当前用户密码，返回有效期 5 分钟的一次性确认 token。
    用于前端在执行高危操作（如删除测评周期/项目）前的二次鉴权。

    请求体::

        { "password": "用户密码" }

    响应::

        { "confirm_token": "uuid-string" }
    """
    permission_classes = [IsAuthenticated]

    def post(self, request):
        password = request.data.get('password', '')
        if not password:
            return Response({'detail': '请输入密码'}, status=status.HTTP_400_BAD_REQUEST)
        if not request.user.check_password(password):
            log_action(
                user=request.user,
                action='verify_password_fail',
                module=OperationLog.MODULE_AUTH,
                level=OperationLog.LEVEL_WARNING,
                target_type='user',
                target_id=request.user.id,
                target_repr=request.user.username,
                is_abnormal=True,
                request=request,
            )
            return Response({'detail': '密码错误'}, status=status.HTTP_401_UNAUTHORIZED)
        token = str(uuid.uuid4())
        cache_key = f'confirm_token:{request.user.id}:{token}'
        cache.set(cache_key, True, timeout=_VERIFY_TOKEN_TTL)
        return Response({'confirm_token': token})


class ChangePasswordView(APIView):
    """
    POST /api/v1/auth/change-password/
    当前登录用户修改自己的密码。

    请求体::

        { "old_password": "旧密码", "new_password": "新密码" }
    """
    permission_classes = [IsAuthenticated]

    def post(self, request):
        old_password = request.data.get('old_password', '')
        new_password = request.data.get('new_password', '')
        if not old_password or not new_password:
            return Response({'detail': '请填写旧密码和新密码'}, status=status.HTTP_400_BAD_REQUEST)
        if len(new_password) < 6:
            return Response({'detail': '新密码长度不能少于 6 位'}, status=status.HTTP_400_BAD_REQUEST)
        if not request.user.check_password(old_password):
            log_action(
                user=request.user,
                action='change_password_fail',
                module=OperationLog.MODULE_AUTH,
                level=OperationLog.LEVEL_WARNING,
                target_type='user',
                target_id=request.user.id,
                target_repr=request.user.username,
                is_abnormal=True,
                request=request,
            )
            return Response({'detail': '旧密码错误'}, status=status.HTTP_401_UNAUTHORIZED)
        request.user.set_password(new_password)
        request.user.must_change_password = False
        request.user.save(update_fields=['password', 'must_change_password'])
        log_action(
            user=request.user,
            action='change_password',
            module=OperationLog.MODULE_AUTH,
            level=OperationLog.LEVEL_NOTICE,
            target_type='user',
            target_id=request.user.id,
            target_repr=request.user.username,
            request=request,
        )
        return Response({'detail': '密码修改成功', 'must_change_password': False})


class CurrentUserView(APIView):
    """GET /api/v1/users/me/ 当前用户信息与角色列表。"""
    permission_classes = [IsAuthenticated]

    def get(self, request):
        serializer = UserSerializer(request.user)
        return Response(serializer.data)


class RoleViewSet(ReadOnlyModelViewSet):
    """GET /api/v1/roles/ 角色列表（只读）。"""
    queryset = Role.objects.all().order_by('level', 'id')
    permission_classes = [IsAuthenticated]
    serializer_class = RoleSerializer


class UserViewSet(ModelViewSet):
    """用户 CRUD：管理员可列表/创建/改删，本人可读/改自己。"""
    permission_classes = [IsAuthenticated]

    @staticmethod
    def _is_framework_admin_account(user_obj):
        """是否为 Django 内置后台账号（is_staff / is_superuser）。"""
        return bool(user_obj.is_staff or user_obj.is_superuser)

    def get_queryset(self):
        # 业务用户管理 API 与 Django 内置后台账号隔离，避免越权与误删风险。
        qs = User.objects.exclude(
            Q(is_staff=True) | Q(is_superuser=True)
        ).prefetch_related('user_roles__role')
        if self.action == 'list':
            qs = qs.order_by('username', 'id')
        else:
            qs = qs.order_by('id')
        if self.action != 'list':
            return qs
        params = self.request.query_params
        if params.get('is_active') not in (None, ''):
            qs = qs.filter(is_active=params.get('is_active').lower() == 'true')
        if params.get('department'):
            qs = qs.filter(department_id=params.get('department'))
        if params.get('class_obj'):
            qs = qs.filter(class_obj_id=params.get('class_obj'))
        if params.get('major'):
            qs = qs.filter(class_obj__major_id=params.get('major'))
        if params.get('grade'):
            qs = qs.filter(class_obj__grade=params.get('grade'))
        if params.get('role'):
            try:
                target_role = Role.objects.get(pk=params.get('role'))
            except Role.DoesNotExist:
                target_role = None
            if target_role is not None:
                qs = qs.annotate(
                    _max_role_level=Max('user_roles__role__level')
                ).filter(_max_role_level=target_role.level)
        search = (params.get('search') or '').strip()
        if search:
            qs = qs.filter(
                Q(username__icontains=search)
                | Q(name__icontains=search)
                | Q(student_no__icontains=search)
                | Q(employee_no__icontains=search)
            )
        return qs

    def get_serializer_class(self):
        if self.action in ('create', 'update', 'partial_update'):
            return UserCreateUpdateSerializer
        if self.action == 'list':
            return UserListSerializer
        return UserSerializer

    def get_permissions(self):
        from .permissions import IsAdminOrReadSelf
        return [IsAdminOrReadSelf()]

    def perform_create(self, serializer):
        instance = serializer.save()
        log_action(
            user=self.request.user,
            action='user_create',
            module=OperationLog.MODULE_USERS,
            level=OperationLog.LEVEL_WARNING,
            target_type='user',
            target_id=instance.id,
            target_repr=instance.username,
            request=self.request,
        )

    def perform_update(self, serializer):
        if self._is_framework_admin_account(serializer.instance):
            from rest_framework.exceptions import PermissionDenied
            raise PermissionDenied('Django 内置后台账号不允许在业务用户管理中修改')
        TRACKED = {
            'phone': '手机号',
            'email': '邮箱',
            'is_active': '账号状态',
            'gender': '性别',
            'student_no': '学号',
            'employee_no': '工号',
            'department_id': '院系',
            'class_obj_id': '班级',
            'name': '姓名',
        }
        old_vals = {f: getattr(serializer.instance, f) for f in TRACKED}
        instance = serializer.save()
        changed = {}
        for field, label in TRACKED.items():
            old_v = old_vals[field]
            new_v = getattr(instance, field)
            if old_v != new_v:
                changed[label] = {
                    'old': str(old_v) if old_v is not None else '',
                    'new': str(new_v) if new_v is not None else '',
                }
        if 'password' in serializer.validated_data:
            changed['密码'] = {'old': '***', 'new': '***（已更新）'}
        log_action(
            user=self.request.user,
            action='user_update',
            module=OperationLog.MODULE_USERS,
            level=OperationLog.LEVEL_WARNING,
            target_type='user',
            target_id=instance.id,
            target_repr=instance.username,
            extra={'changed': changed} if changed else {},
            request=self.request,
        )

    def perform_destroy(self, instance):
        if self._is_framework_admin_account(instance):
            from rest_framework.exceptions import PermissionDenied
            raise PermissionDenied('Django 内置后台账号不允许在业务用户管理中删除')
        target_repr = instance.username
        target_id = instance.id
        instance.delete()
        log_action(
            user=self.request.user,
            action='user_delete',
            module=OperationLog.MODULE_USERS,
            level=OperationLog.LEVEL_CRITICAL,
            target_type='user',
            target_id=target_id,
            target_repr=target_repr,
            is_audit_event=True,
            request=self.request,
        )

    def _parse_batch_user_ids(self, request):
        """解析并校验批量用户 ID 列表。"""
        user_ids = request.data.get('user_ids')
        if not isinstance(user_ids, list) or not user_ids:
            return None, Response(
                {'detail': 'user_ids 必须为非空数组'},
                status=status.HTTP_400_BAD_REQUEST
            )
        parsed_ids = []
        for uid in user_ids:
            try:
                parsed_ids.append(int(uid))
            except (TypeError, ValueError):
                return None, Response(
                    {'detail': 'user_ids 中存在非法 ID'},
                    status=status.HTTP_400_BAD_REQUEST
                )
        if request.user.id in parsed_ids:
            return None, Response(
                {'detail': '不允许对当前登录账号执行批量操作'},
                status=status.HTTP_400_BAD_REQUEST
            )
        users_qs = User.objects.filter(id__in=parsed_ids)
        found_ids = set(users_qs.values_list('id', flat=True))
        missing = [i for i in parsed_ids if i not in found_ids]
        if missing:
            return None, Response(
                {'detail': f'部分用户不存在: {missing}'},
                status=status.HTTP_400_BAD_REQUEST
            )
        has_superadmin = users_qs.filter(
            Q(current_role__level__gte=5) | Q(user_roles__role__level__gte=5)
        ).distinct().exists()
        if has_superadmin:
            return None, Response(
                {'detail': f'批量操作不允许包含{get_role_display_name(ROLE_LEVEL_SUPERADMIN)}账号'},
                status=status.HTTP_400_BAD_REQUEST
            )
        has_framework_admin = users_qs.filter(Q(is_staff=True) | Q(is_superuser=True)).exists()
        if has_framework_admin:
            return None, Response(
                {'detail': '批量操作不允许包含 Django 内置后台账号'},
                status=status.HTTP_400_BAD_REQUEST
            )
        return users_qs, None

    @action(detail=False, methods=['post'], url_path='batch/set-active')
    def batch_set_active(self, request):
        """批量启用/禁用用户。"""
        from .permissions import user_is_super_admin
        if not user_is_super_admin(request.user):
            return Response({'detail': f'仅{get_role_display_name(ROLE_LEVEL_SUPERADMIN)}可执行批量操作'}, status=status.HTTP_403_FORBIDDEN)
        is_active = request.data.get('is_active')
        if not isinstance(is_active, bool):
            return Response({'detail': 'is_active 必须为布尔值'}, status=status.HTTP_400_BAD_REQUEST)
        users_qs, err = self._parse_batch_user_ids(request)
        if err:
            return err
        affected = users_qs.update(is_active=is_active)
        from realtime.registry import broadcast
        broadcast({'type': 'data_changed', 'model': 'user'})
        log_action(
            user=request.user,
            action='user_batch_set_active',
            module=OperationLog.MODULE_USERS,
            level=OperationLog.LEVEL_WARNING,
            target_type='user',
            target_id=0,
            target_repr='batch',
            extra={
                'is_active': is_active,
                'affected_count': affected,
                'user_ids': list(users_qs.values_list('id', flat=True)),
            },
            is_audit_event=True,
            request=request,
        )
        return Response({'affected_count': affected}, status=status.HTTP_200_OK)

    @action(detail=False, methods=['post'], url_path='batch/reset-password')
    def batch_reset_password(self, request):
        """批量重置密码。"""
        from .permissions import user_is_super_admin
        if not user_is_super_admin(request.user):
            return Response({'detail': f'仅{get_role_display_name(ROLE_LEVEL_SUPERADMIN)}可执行批量操作'}, status=status.HTTP_403_FORBIDDEN)
        new_password = (request.data.get('new_password') or '').strip()
        if len(new_password) < 6:
            return Response({'detail': 'new_password 长度至少为 6'}, status=status.HTTP_400_BAD_REQUEST)
        users_qs, err = self._parse_batch_user_ids(request)
        if err:
            return err
        for user in users_qs:
            user.set_password(new_password)
            user.save(update_fields=['password'])
        affected = users_qs.count()
        log_action(
            user=request.user,
            action='user_batch_reset_password',
            module=OperationLog.MODULE_USERS,
            level=OperationLog.LEVEL_CRITICAL,
            target_type='user',
            target_id=0,
            target_repr='batch',
            extra={
                'affected_count': affected,
                'user_ids': list(users_qs.values_list('id', flat=True)),
            },
            is_audit_event=True,
            request=request,
        )
        return Response({'affected_count': affected}, status=status.HTTP_200_OK)

    @action(detail=False, methods=['post'], url_path='batch/set-role')
    def batch_set_role(self, request):
        """批量重设角色（同一套角色应用到所选用户）。"""
        from .permissions import user_is_super_admin
        if not user_is_super_admin(request.user):
            return Response({'detail': f'仅{get_role_display_name(ROLE_LEVEL_SUPERADMIN)}可执行批量操作'}, status=status.HTTP_403_FORBIDDEN)
        role_ids = request.data.get('role_ids')
        responsible_class_ids = request.data.get('responsible_class_ids', [])
        if not isinstance(role_ids, list) or not role_ids:
            return Response({'detail': 'role_ids 必须为非空数组'}, status=status.HTTP_400_BAD_REQUEST)
        if not isinstance(responsible_class_ids, list):
            return Response({'detail': 'responsible_class_ids 必须为数组'}, status=status.HTTP_400_BAD_REQUEST)
        role_ids = list(Role.objects.filter(level__lt=ROLE_LEVEL_SUPERADMIN, id__in=role_ids).values_list('id', flat=True))
        if not role_ids:
            return Response({'detail': '可分配角色不能为空'}, status=status.HTTP_400_BAD_REQUEST)
        users_qs, err = self._parse_batch_user_ids(request)
        if err:
            return err

        # 复用用户编辑序列化器中的角色同步逻辑，保证单个与批量行为一致。
        sync_serializer = UserCreateUpdateSerializer()
        for user in users_qs:
            sync_serializer._sync_user_roles(user, role_ids, responsible_class_ids)
        affected = users_qs.count()
        log_action(
            user=request.user,
            action='user_batch_set_role',
            module=OperationLog.MODULE_USERS,
            level=OperationLog.LEVEL_WARNING,
            target_type='user',
            target_id=0,
            target_repr='batch',
            extra={
                'affected_count': affected,
                'role_ids': role_ids,
                'responsible_class_ids': responsible_class_ids,
                'user_ids': list(users_qs.values_list('id', flat=True)),
            },
            is_audit_event=True,
            request=request,
        )
        return Response({'affected_count': affected}, status=status.HTTP_200_OK)

    @action(detail=False, methods=['post'], url_path='batch/delete')
    def batch_delete(self, request):
        """批量删除用户。"""
        from .permissions import user_is_super_admin
        if not user_is_super_admin(request.user):
            return Response({'detail': f'仅{get_role_display_name(ROLE_LEVEL_SUPERADMIN)}可执行批量操作'}, status=status.HTTP_403_FORBIDDEN)
        users_qs, err = self._parse_batch_user_ids(request)
        if err:
            return err
        user_ids = list(users_qs.values_list('id', flat=True))
        usernames = list(users_qs.values_list('username', flat=True))
        affected = len(user_ids)
        users_qs.delete()
        log_action(
            user=request.user,
            action='user_batch_delete',
            module=OperationLog.MODULE_USERS,
            level=OperationLog.LEVEL_CRITICAL,
            target_type='user',
            target_id=0,
            target_repr='batch',
            extra={
                'affected_count': affected,
                'user_ids': user_ids,
                'usernames': usernames,
            },
            is_audit_event=True,
            request=request,
        )
        return Response({'affected_count': affected}, status=status.HTTP_200_OK)


class UserImportTemplateAPIView(APIView):
    """
    GET /api/v1/users/import/template/ 生成 Excel 导入用户模板（树形分表）。

    按角色拆分为独立 sheet，处理顺序：学生 → 学生助理 → 评审老师 → 院系主任。
    各层只需声明本级属性 + 下级归属键，系统自动沿树形向上汇总 scope。
    """
    permission_classes = [IsAuthenticated]

    def get(self, request):
        """
        生成树形分表导入模板，按操作人权限控制可导入的角色 sheet。
        """
        from users.permissions import user_level_at_least, get_user_level
        if not user_level_at_least(request.user, 2):
            return Response({'detail': f'{get_role_display_name(ROLE_LEVEL_COUNSELOR)}及以上角色可下载导入模板'}, status=status.HTTP_403_FORBIDDEN)
        operator_level = get_user_level(request.user)
        is_super = operator_level >= 5
        is_director = operator_level == 3
        is_counselor = operator_level == 2

        import openpyxl
        from openpyxl.styles import Font, PatternFill, Alignment
        from django.http import HttpResponse
        from io import BytesIO

        all_roles = list(Role.objects.all().order_by('level'))
        role_name_map = {r.code: r.name for r in all_roles}
        student_label = role_name_map.get('student', '学生')
        assistant_label = role_name_map.get('student_assistant', get_role_display_name(ROLE_LEVEL_ASSISTANT))
        counselor_label = role_name_map.get('counselor', get_role_display_name(ROLE_LEVEL_COUNSELOR))
        director_label = role_name_map.get('director', get_role_display_name(ROLE_LEVEL_DIRECTOR))

        wb = openpyxl.Workbook()
        header_font = Font(bold=True)
        header_fill = PatternFill(start_color='DCE6F1', end_color='DCE6F1', fill_type='solid')

        def _style_header(ws, col_count):
            """@private 为表头行设置统一样式。"""
            for col in range(1, col_count + 1):
                cell = ws.cell(row=1, column=col)
                cell.font = header_font
                cell.fill = header_fill
                cell.alignment = Alignment(horizontal='center')

        # ── Sheet 1: 学生导入 ──
        ws_stu = wb.active
        ws_stu.title = f'{student_label}导入'
        stu_headers = ['用户名*', '学号*', '姓名', '性别', '邮箱', '手机号',
                       '院系名称', '班级名称', '专业名称', '年级', '初始密码']
        ws_stu.append(stu_headers)
        ws_stu.append(['stu001', '2021001', '张三', '男', 'zhangsan@example.com',
                       '13800138000', '计算机学院', '软件1班', '软件工程', '2021', '123456'])
        _style_header(ws_stu, len(stu_headers))

        # ── Sheet 2: 学生助理导入（从学生中提拔，仅填学号） ──
        ws_asst = wb.create_sheet(f'{assistant_label}导入')
        asst_headers = ['学号*']
        ws_asst.append(asst_headers)
        ws_asst.append(['2021001'])
        _style_header(ws_asst, len(asst_headers))

        # ── Sheet 3: 评审老师导入（需院系主任或超管权限） ──
        if is_super or is_director:
            ws_coun = wb.create_sheet(f'{counselor_label}导入')
            coun_headers = ['用户名*', '工号*', '姓名', '性别', '邮箱', '手机号',
                            '院系名称*', '负责专业列表', '负责班级列表', '初始密码']
            ws_coun.append(coun_headers)
            ws_coun.append(['teacher01', 'T001', '李老师', '男', 'li@example.com',
                            '13900139000', '计算机学院', '软件工程;计算机科学', '', '123456'])
            _style_header(ws_coun, len(coun_headers))

        # ── Sheet 4: 院系主任导入（仅超管可导入） ──
        if is_super:
            ws_dir = wb.create_sheet(f'{director_label}导入')
            dir_headers = ['用户名*', '工号*', '姓名', '性别', '邮箱', '手机号',
                           '负责院系名称*', '初始密码']
            ws_dir.append(dir_headers)
            ws_dir.append(['director01', 'D001', '王主任', '男', 'wang@example.com',
                           '13600136000', '计算机学院', '123456'])
            _style_header(ws_dir, len(dir_headers))

        # ── 辅助 Sheet: 导入前置检查 ──
        ws_pre = wb.create_sheet('导入前置检查')
        ws_pre.append(['检查项', '是否必须', '说明'])
        ws_pre.append(['院系主数据', '必须', '导入前请先在系统中维护院系，导入表中的院系名称需完全匹配'])
        ws_pre.append(['班级主数据', '必须', '导入前请先在系统中维护班级，导入表中的班级名称+院系(+年级)需匹配'])
        ws_pre.append(['专业主数据', '建议', '若评审老师使用"负责专业列表"，专业须已存在且下面有班级'])
        ws_pre.append(['分表结构', '说明', '每张工作表对应一种角色，只需填写需要导入的表，无数据的表自动跳过'])
        ws_pre.append(['处理顺序', '自动',
                       f'系统按 {student_label}→{assistant_label}→{counselor_label}→{director_label} 顺序处理'])
        ws_pre.append(['预检建议', '强烈建议', '先点「预检」查看错误与依赖关系摘要，再执行正式导入'])
        ws_pre.append(['风险提示', '重点', '若院系/班级/专业尚未创建，将出现"不存在"错误；可在预检后按权限一键创建'])
        red_font = Font(color='FF0000', bold=True)
        for row_idx in (2, 3):
            for col_idx in (1, 2, 3):
                ws_pre.cell(row=row_idx, column=col_idx).font = red_font

        # ── 辅助 Sheet: 树形导入规则 ──
        ws_tree = wb.create_sheet('树形导入规则')
        ws_tree.append(['层级', '工作表', '关键字段', '归属规则'])
        ws_tree.append(['1（底层）', f'{student_label}导入', '院系名称 + 班级名称',
                        '学生归属到班级，是树的叶子节点'])
        ws_tree.append(['2', f'{assistant_label}导入', '学号',
                        '从已有学生中提拔，仅填学号；默认负责自身班级，跨班由评审规则控制'])
        ws_tree.append(['3', f'{counselor_label}导入', '院系名称 + 负责专业/班级列表',
                        '管辖指定班级的学生；负责专业列表会自动展开为该专业下所有班级'])
        ws_tree.append(['4（顶层）', f'{director_label}导入', '负责院系名称',
                        f'自动继承该院系下所有{counselor_label}的负责班级，无需手动指定'])
        ws_tree.append(['', '', '', ''])
        ws_tree.append(['树形汇总', '', '',
                        f'{director_label} ←(按院系)← {counselor_label} ←(按班级)← {student_label}'])
        if is_counselor:
            ws_tree.append(['当前账号限制', '', '仅负责班级', '超出负责班级范围将被拒绝'])
        elif is_director:
            ws_tree.append(['当前账号限制', '', '仅本院系', '跨院系将被拒绝'])

        # ── 辅助 Sheet: 各表字段说明 ──
        ws_field = wb.create_sheet('字段说明')
        ws_field.append(['工作表', '字段', '必填', '说明'])
        ws_field.append([f'{student_label}导入', '用户名', '是', '登录用户名，全局唯一'])
        ws_field.append([f'{student_label}导入', '学号', '是', '学号，全局唯一'])
        ws_field.append([f'{student_label}导入', '姓名/性别/邮箱/手机号', '否', '基本信息'])
        ws_field.append([f'{student_label}导入', '院系名称', '建议', '需与系统中院系名称完全一致'])
        ws_field.append([f'{student_label}导入', '班级名称', '建议', '需与系统中班级名称完全一致'])
        ws_field.append([f'{student_label}导入', '专业名称', '否', '系统按 院系+专业 匹配'])
        ws_field.append([f'{student_label}导入', '年级', '否', '如 2021、2022'])
        ws_field.append([f'{student_label}导入', '初始密码', '否', '留空默认 123456'])
        ws_field.append(['', '', '', ''])
        ws_field.append([f'{assistant_label}导入', '学号', '是',
                         '已存在的学生学号（本文件学生表或数据库中）；默认负责自身所在班级，跨班分配由评审规则控制'])
        ws_field.append(['', '', '', ''])
        ws_field.append([f'{counselor_label}导入', '用户名/工号', '是', '登录用户名与工号，全局唯一'])
        ws_field.append([f'{counselor_label}导入', '姓名/性别/邮箱/手机号', '否', '基本信息'])
        ws_field.append([f'{counselor_label}导入', '院系名称', '是', '所属院系（不可跨院系）'])
        ws_field.append([f'{counselor_label}导入', '负责专业列表', '否',
                         '展开为该院系该专业下所有班级；多个用逗号/分号分隔'])
        ws_field.append([f'{counselor_label}导入', '负责班级列表', '否',
                         '显式指定班级名称；可与专业列表合并去重'])
        ws_field.append([f'{counselor_label}导入', '初始密码', '否', '留空默认 123456'])
        ws_field.append(['', '', '', ''])
        ws_field.append([f'{director_label}导入', '用户名/工号', '是', '登录用户名与工号，全局唯一'])
        ws_field.append([f'{director_label}导入', '姓名/性别/邮箱/手机号', '否', '基本信息'])
        ws_field.append([f'{director_label}导入', '负责院系名称', '是',
                         '可多个逗号分隔；自动继承该院系下所有评审老师的负责班级'])
        ws_field.append([f'{director_label}导入', '初始密码', '否', '留空默认 123456'])
        ws_field.append(['', '', '', ''])
        if is_director:
            ws_field.append(['导入范围（当前账号）', '', '—',
                             f'仅本院系；可导入{student_label}、{assistant_label}、{counselor_label}'])
        elif is_counselor:
            ws_field.append(['导入范围（当前账号）', '', '—',
                             f'仅负责班级；可导入{student_label}、{assistant_label}'])
        else:
            ws_field.append(['导入范围（当前账号）', '', '—', f'全局范围；不可导入{get_role_display_name(ROLE_LEVEL_SUPERADMIN)}'])

        # ── 辅助 Sheet: 角色代码说明 ──
        ws_role = wb.create_sheet('角色代码说明')
        ws_role.append(['角色代码', '角色名称', '等级', '对应工作表', '说明'])
        if is_super:
            importable_codes = {'student', 'student_assistant', 'counselor', 'director'}
        elif is_director:
            importable_codes = {'student', 'student_assistant', 'counselor'}
        else:
            importable_codes = {'student', 'student_assistant'}
        sheet_map = {
            'student': f'{student_label}导入', 'student_assistant': f'{assistant_label}导入',
            'counselor': f'{counselor_label}导入', 'director': f'{director_label}导入',
        }
        for r in all_roles:
            if r.code in importable_codes:
                ws_role.append([r.code, r.name, str(r.level), sheet_map.get(r.code, ''), r.description or ''])

        # ── 辅助 Sheet: 未知项处理说明 ──
        ws_fix = wb.create_sheet('未知项处理说明')
        ws_fix.append(['步骤', '操作', '说明'])
        ws_fix.append(['1', '上传并点击「预检」', '系统返回各表的错误清单 + 依赖关系摘要 + 未知组织项列表'])
        ws_fix.append(['2', '处理未知组织项', '可选择"映射到已有项"或"按权限一键创建缺失项"'])
        ws_fix.append(['3', '再次预检', '确认错误清零或仅保留可接受告警'])
        ws_fix.append(['4', '正式导入', '点击「确认导入」正式写入数据库'])

        buf = BytesIO()
        wb.save(buf)
        buf.seek(0)
        resp = HttpResponse(buf.getvalue(), content_type='application/vnd.openxmlformats-officedocument.spreadsheetml.sheet')
        resp['Content-Disposition'] = 'attachment; filename="user_import_template.xlsx"'
        return resp


class UserImportAPIView(APIView):
    """
    POST /api/v1/users/import/ Excel 批量导入用户。

    权限分级：
    - LV5 超级管理员：无范围限制，可导入任意角色（除 superadmin）
    - LV3 院系主任：只能导入本院系用户，导入角色等级不超过 LV2
    - LV2 评审老师（辅导员）：只能导入负责班级用户，导入角色限定为学生/学生助理（LV0/LV1）

    安全控制：
    - 非超级管理员导入的用户默认 is_active=False，需上级审批激活
    - 所有导入操作均记录 WARNING 级操作日志
    - 导入角色等级不得 >= 操作人等级（防提权）
    """
    permission_classes = [IsAuthenticated]
    parser_classes = [MultiPartParser, FormParser]

    @staticmethod
    def _is_truthy(value):
        """解析 form-data 布尔值。"""
        return str(value).strip().lower() in ('1', 'true', 'yes', 'on')

    @staticmethod
    def _preview_cache_key(user_id, token):
        """构造导入预检缓存键。"""
        return f'import_preview:{user_id}:{token}'

    @staticmethod
    def _split_multi_values(raw_value):
        """将逗号/分号/顿号分隔文本解析为字符串列表。"""
        if raw_value is None:
            return []
        text = str(raw_value).strip()
        if not text:
            return []
        for sep in ['；', ';', '，', ',', '、', '\n', '\r']:
            text = text.replace(sep, ',')
        vals = []
        for item in text.split(','):
            cur = item.strip()
            if cur and cur not in vals:
                vals.append(cur)
        return vals

    @staticmethod
    def _normalize_header(text):
        """统一表头命名格式，便于兼容不同模板列名。"""
        if text is None:
            return ''
        return str(text).strip().replace('*', '').replace('（', '(').replace('）', ')')

    @staticmethod
    def _get_sheet_role_map():
        """动态构建 sheet 名称到 role_code 的映射，兼容默认名称和自定义角色名。"""
        from users.role_resolver import get_role_display_name
        mapping = {}
        code_level_pairs = [
            ('student', 0), ('student_assistant', 1), ('counselor', 2), ('director', 3),
        ]
        defaults = {
            0: get_role_display_name(0),
            1: get_role_display_name(1),
            2: get_role_display_name(2),
            3: get_role_display_name(3),
        }
        for code, level in code_level_pairs:
            name = get_role_display_name(level, defaults[level])
            mapping[f'{name}导入'] = code
            default_name = defaults[level]
            if name != default_name:
                mapping[f'{default_name}导入'] = code
        return mapping

    # 按树形依赖排序的优先级（值越小越先处理）
    _ROLE_PRIORITY = {'student': 0, 'student_assistant': 1, 'counselor': 2, 'director': 3}

    def _load_rows_from_excel(self, file_obj):
        """
        读取 Excel 并转换为统一行结构（支持树形分表和旧版单表）。

        新模板：按 sheet 名称自动识别角色，各 sheet 有独立列结构。
        旧模板：回退读取 active sheet，依赖 role_code 列。

        @returns {list[dict]} rows - 统一格式行列表，含 source_sheet / role_code 字段
        """
        import openpyxl

        wb = openpyxl.load_workbook(file_obj, read_only=True)

        sheet_role_map = self._get_sheet_role_map()
        matched_sheets = []
        for ws in wb.worksheets:
            title = (ws.title or '').strip()
            if title in sheet_role_map:
                matched_sheets.append((ws, sheet_role_map[title]))

        if matched_sheets:
            return self._parse_multi_sheet(matched_sheets)

        # 向后兼容：旧版单表模板
        return self._parse_legacy_sheet(wb.active, source_sheet='')

    def _parse_legacy_sheet(self, ws, source_sheet=''):
        """@private 解析旧版单一 sheet 模板（含 role_code 列）。"""
        header_row = next(ws.iter_rows(min_row=1, max_row=1, values_only=True), None) or []
        headers = [self._normalize_header(h) for h in header_row]
        idx_map = {h: i for i, h in enumerate(headers) if h}

        aliases = {
            'username': ['用户名', '账号'],
            'student_no': ['学号'],
            'employee_no': ['工号'],
            'full_name': ['姓名'],
            'gender': ['性别'],
            'email': ['邮箱'],
            'phone': ['手机号', '手机'],
            'department_name': ['院系名称', '院系'],
            'class_name': ['班级名称', '班级'],
            'major_name': ['专业名称', '专业'],
            'grade': ['年级'],
            'role_code': ['角色代码', '角色'],
            'responsible_classes': ['负责班级列表'],
            'managed_counselors': ['管理评审老师用户名列表'],
            'password': ['初始密码', '密码'],
        }

        def pick(row, key):
            for alias in aliases.get(key, []):
                idx = idx_map.get(alias)
                if idx is not None and idx < len(row):
                    return row[idx]
            return None

        rows = []
        for row_idx, row in enumerate(ws.iter_rows(min_row=2, values_only=True), start=2):
            if not row or all((c is None or str(c).strip() == '') for c in row):
                continue
            rows.append({
                'row_num': row_idx,
                'source_sheet': source_sheet,
                'username': str(pick(row, 'username') or '').strip(),
                'student_no': str(pick(row, 'student_no') or '').strip(),
                'employee_no': str(pick(row, 'employee_no') or '').strip(),
                'full_name': str(pick(row, 'full_name') or '').strip(),
                'gender': str(pick(row, 'gender') or '').strip(),
                'email': str(pick(row, 'email') or '').strip(),
                'phone': str(pick(row, 'phone') or '').strip(),
                'department_name': str(pick(row, 'department_name') or '').strip(),
                'class_name': str(pick(row, 'class_name') or '').strip(),
                'major_name': str(pick(row, 'major_name') or '').strip(),
                'grade': str(pick(row, 'grade') or '').strip(),
                'role_code': str(pick(row, 'role_code') or '').strip(),
                'responsible_classes': str(pick(row, 'responsible_classes') or '').strip(),
                'responsible_majors': '',
                'responsible_departments': '',
                'managed_counselors': str(pick(row, 'managed_counselors') or '').strip(),
                'password': str(pick(row, 'password') or '').strip(),
            })
        return rows

    def _parse_multi_sheet(self, matched_sheets):
        """
        @private 解析树形分表模板，各 sheet 按角色拥有独立列结构。

        @param {list[tuple]} matched_sheets - [(worksheet, role_code), ...]
        @returns {list[dict]} 统一格式行列表
        """
        all_rows = []
        for ws, role_code in matched_sheets:
            header_row = next(ws.iter_rows(min_row=1, max_row=1, values_only=True), None) or []
            headers = [self._normalize_header(h) for h in header_row]
            idx_map = {h: i for i, h in enumerate(headers) if h}
            sheet_title = ws.title or ''

            def _pick(row, *col_names):
                """@private 按列名别名取值。"""
                for name in col_names:
                    idx = idx_map.get(name)
                    if idx is not None and idx < len(row):
                        val = row[idx]
                        if val is not None:
                            return str(val).strip()
                return ''

            for row_idx, row in enumerate(ws.iter_rows(min_row=2, values_only=True), start=2):
                if not row or all((c is None or str(c).strip() == '') for c in row):
                    continue

                if role_code == 'student':
                    all_rows.append({
                        'row_num': row_idx, 'source_sheet': sheet_title, 'role_code': 'student',
                        'username': _pick(row, '用户名', '账号'),
                        'student_no': _pick(row, '学号'),
                        'employee_no': '',
                        'full_name': _pick(row, '姓名'),
                        'gender': _pick(row, '性别'),
                        'email': _pick(row, '邮箱'),
                        'phone': _pick(row, '手机号', '手机'),
                        'department_name': _pick(row, '院系名称', '院系'),
                        'class_name': _pick(row, '班级名称', '班级'),
                        'major_name': _pick(row, '专业名称', '专业'),
                        'grade': _pick(row, '年级'),
                        'responsible_classes': '',
                        'responsible_majors': '',
                        'responsible_departments': '',
                        'managed_counselors': '',
                        'password': _pick(row, '初始密码', '密码'),
                    })

                elif role_code == 'student_assistant':
                    all_rows.append({
                        'row_num': row_idx, 'source_sheet': sheet_title, 'role_code': 'student_assistant',
                        'username': '', 'employee_no': '', 'full_name': '', 'gender': '',
                        'email': '', 'phone': '', 'department_name': '', 'class_name': '',
                        'major_name': '', 'grade': '', 'managed_counselors': '', 'password': '',
                        'responsible_departments': '',
                        'student_no': _pick(row, '学号'),
                        'responsible_classes': '',
                        'responsible_majors': '',
                    })

                elif role_code == 'counselor':
                    all_rows.append({
                        'row_num': row_idx, 'source_sheet': sheet_title, 'role_code': 'counselor',
                        'username': _pick(row, '用户名', '账号'),
                        'employee_no': _pick(row, '工号'),
                        'student_no': '',
                        'full_name': _pick(row, '姓名'),
                        'gender': _pick(row, '性别'),
                        'email': _pick(row, '邮箱'),
                        'phone': _pick(row, '手机号', '手机'),
                        'department_name': _pick(row, '院系名称', '院系'),
                        'class_name': '', 'major_name': '', 'grade': '',
                        'managed_counselors': '',
                        'responsible_departments': '',
                        'responsible_majors': _pick(row, '负责专业列表'),
                        'responsible_classes': _pick(row, '负责班级列表'),
                        'password': _pick(row, '初始密码', '密码'),
                    })

                elif role_code == 'director':
                    all_rows.append({
                        'row_num': row_idx, 'source_sheet': sheet_title, 'role_code': 'director',
                        'username': _pick(row, '用户名', '账号'),
                        'employee_no': _pick(row, '工号'),
                        'student_no': '',
                        'full_name': _pick(row, '姓名'),
                        'gender': _pick(row, '性别'),
                        'email': _pick(row, '邮箱'),
                        'phone': _pick(row, '手机号', '手机'),
                        'department_name': '',
                        'class_name': '', 'major_name': '', 'grade': '',
                        'managed_counselors': '',
                        'responsible_classes': '',
                        'responsible_majors': '',
                        'responsible_departments': _pick(row, '负责院系名称', '院系名称', '院系'),
                        'password': _pick(row, '初始密码', '密码'),
                    })

        # 按树形依赖顺序排序
        all_rows.sort(key=lambda r: self._ROLE_PRIORITY.get(r.get('role_code', ''), 99))
        return all_rows

    def _operator_scope(self, request):
        """计算操作者分级权限范围。"""
        from users.permissions import user_level_at_least, get_user_level
        from org.models import Class
        from .models import UserRole as UserRoleModel

        operator_level = get_user_level(request.user)
        if not user_level_at_least(request.user, 2):
            return None, Response({'detail': f'{get_role_display_name(ROLE_LEVEL_COUNSELOR)}及以上角色可导入用户'}, status=status.HTTP_403_FORBIDDEN)

        is_super = operator_level >= 5
        is_director = operator_level == 3
        is_counselor = operator_level == 2

        allowed_role_codes = None
        if is_director and not is_super:
            allowed_role_codes = {'student', 'student_assistant', 'counselor'}
        if is_counselor and not is_super:
            allowed_role_codes = {'student', 'student_assistant'}

        allowed_dept_ids = None
        if is_director and not is_super:
            if request.user.department_id:
                allowed_dept_ids = {request.user.department_id}
            else:
                return None, Response({'detail': '您未关联院系，无法限定导入范围'}, status=status.HTTP_403_FORBIDDEN)

        allowed_class_ids = None
        counselor_managed_dept_ids = set()
        if is_counselor and not is_super:
            scope_ids = set(
                UserRoleModel.objects.filter(
                    user=request.user, scope_type='class', scope_id__isnull=False
                ).values_list('scope_id', flat=True)
            )
            if not scope_ids:
                return None, Response({'detail': '您未关联任何班级，无法导入'}, status=status.HTTP_403_FORBIDDEN)
            allowed_class_ids = scope_ids
            counselor_managed_dept_ids = set(
                Class.objects.filter(id__in=scope_ids).values_list('department_id', flat=True)
            )

        return {
            'operator_level': operator_level,
            'is_super': is_super,
            'is_director': is_director,
            'is_counselor': is_counselor,
            'allowed_role_codes': allowed_role_codes,
            'allowed_dept_ids': allowed_dept_ids,
            'allowed_class_ids': allowed_class_ids,
            'counselor_managed_dept_ids': counselor_managed_dept_ids,
        }, None

    def _parse_resolution_payload(self, request):
        """解析前端传来的修复/映射规则。"""
        payload = request.data.get('resolution_payload')
        if payload in (None, ''):
            return {}
        if isinstance(payload, dict):
            return payload
        if isinstance(payload, str):
            try:
                return json.loads(payload)
            except json.JSONDecodeError:
                return {}
        return {}

    def _ensure_major_grade(self, major_obj, grade_text):
        """确保专业 grades 中包含年级。"""
        if not major_obj or not grade_text:
            return
        grades = list(major_obj.grades or [])
        if grade_text not in grades:
            grades.append(grade_text)
            major_obj.grades = grades
            major_obj.save(update_fields=['grades', 'updated_at'])

    def _can_create_department(self, scope):
        """是否允许创建院系。安全收口：仅超管可创建院系。"""
        return scope['is_super']

    def _can_create_major(self, scope, dept_id):
        """是否允许在指定院系创建专业。"""
        if scope['is_super']:
            return True
        if scope['is_director']:
            return dept_id in (scope['allowed_dept_ids'] or set())
        if scope['is_counselor']:
            return dept_id in (scope['counselor_managed_dept_ids'] or set())
        return False

    def _can_create_class(self, scope, dept_id):
        """是否允许在指定院系创建班级。"""
        return self._can_create_major(scope, dept_id)

    def _run_import(self, request, rows, scope, dry_run=False, resolution=None,
                    auto_create_missing=False, hash_iterations=None,
                    force_change_password=True, progress_callback=None):
        """
        执行导入核心逻辑（预检/正式导入共用，支持树形分表）。

        行已按 student → student_assistant → counselor → director 排序。
        学生助理为提拔（不新建用户），评审老师支持专业展开，院系主任按院系自动汇总。

        @param {int|None} hash_iterations - 密码哈希迭代次数，None 使用默认
        @param {bool} force_change_password - 是否要求用户首次登录后强制修改密码
        @param {callable|None} progress_callback - fn(current, total) 进度回调
        @returns {dict} 含 row_count, success_count, errors, warnings, unknown_entities, dependency_graph, created_entities
        """
        from org.models import Department, Major, Class
        from users.serializers import UserCreateUpdateSerializer

        from django.contrib.auth.hashers import make_password as _make_password

        resolution = resolution or {}
        department_map = resolution.get('department_map', {}) or {}
        major_map = resolution.get('major_map', {}) or {}
        class_map = resolution.get('class_map', {}) or {}

        use_low_iter = hash_iterations is not None and hash_iterations > 0
        if use_low_iter:
            from users.hashers import LowIterPBKDF2PasswordHasher
            _low_hasher = LowIterPBKDF2PasswordHasher()
            _low_hasher.iterations = hash_iterations

        def _hash_password(raw_password):
            """@private 根据配置选择 hasher 生成密码哈希。"""
            if use_low_iter:
                return _make_password(raw_password, hasher=_low_hasher)
            return _make_password(raw_password)

        processed_count = 0
        total_rows = len(rows)

        errors = []
        warnings = []
        success_count = 0
        created_entities = {'departments': [], 'majors': [], 'classes': []}
        dependency_graph = {
            'supports_mixed_order': True,
            'preconditions': [
                '系统按 学生→学生助理→评审老师→院系主任 顺序处理',
                '院系/专业/班级可在预检后映射或按权限创建',
            ],
            'nodes': {'student': 0, 'student_assistant': 0, 'counselor': 0, 'director': 0},
            'edges': {
                'student_to_class': 0,
                'assistant_to_class': 0,
                'counselor_to_class': 0,
                'director_to_dept': 0,
            },
            'samples': [],
        }
        unknown = {'departments': {}, 'majors': {}, 'classes': {}, 'grades': {}}

        seen_usernames = set()
        seen_student_no = set()
        seen_employee_no = set()
        sync_serializer = UserCreateUpdateSerializer()

        # 同文件导入的学生（学号 → User），供学生助理表引用
        imported_students = {}
        # 同文件导入的评审老师（dept_id → [user]），供院系主任表引用
        imported_counselors_by_dept = {}

        gender_map = {'男': 'M', '女': 'F', '未知': '', '': ''}

        def _row_label(row):
            """@private 生成行定位标签，含 sheet 来源。"""
            sheet = row.get('source_sheet', '')
            num = row.get('row_num', '?')
            return f'[{sheet}!第{num}行]' if sheet else f'[第{num}行]'

        def _err(row, message):
            """@private 添加错误，自动附带 sheet 信息。"""
            errors.append({
                'row': row.get('row_num'),
                'sheet': row.get('source_sheet', ''),
                'message': f'{_row_label(row)} {message}',
            })

        def _warn(row, message):
            """@private 添加警告，自动附带 sheet 信息。"""
            warnings.append({
                'row': row.get('row_num'),
                'sheet': row.get('source_sheet', ''),
                'message': f'{_row_label(row)} {message}',
            })

        def add_unknown(kind, key, payload):
            item = unknown[kind].setdefault(key, {'rows': set(), **payload})
            item['rows'].add(payload['row_num'])

        def resolve_department(dept_name, row_num):
            if not dept_name:
                return None
            mapped_name = str(department_map.get(dept_name, dept_name)).strip()
            dept = Department.objects.filter(name=mapped_name).first()
            if dept:
                return dept
            add_unknown('departments', mapped_name, {'row_num': row_num, 'name': mapped_name})
            if dry_run or not auto_create_missing:
                return None
            if not self._can_create_department(scope):
                errors.append({'row': row_num, 'sheet': '', 'message': f'无权限创建新院系: {mapped_name}'})
                return None
            dept = Department.objects.create(name=mapped_name, code=f'auto_{uuid.uuid4().hex[:8]}')
            created_entities['departments'].append({'id': dept.id, 'name': dept.name})
            return dept

        def resolve_major(department, major_name, row_num):
            if not major_name or not department:
                return None
            map_key = f'{department.name}|{major_name}'
            mapped_name = str(major_map.get(map_key, major_name)).strip()
            major = Major.objects.filter(name=mapped_name, department=department).first()
            if major:
                return major
            add_unknown('majors', f'{department.name}|{mapped_name}', {
                'row_num': row_num, 'department_name': department.name, 'major_name': mapped_name,
            })
            if dry_run or not auto_create_missing:
                return None
            if not self._can_create_major(scope, department.id):
                errors.append({'row': row_num, 'sheet': '', 'message': f'无权限在院系[{department.name}]创建专业: {mapped_name}'})
                return None
            major = Major.objects.create(name=mapped_name, department=department)
            created_entities['majors'].append({'id': major.id, 'name': major.name, 'department': department.name})
            return major

        def resolve_class(department, major_obj, class_name, grade_text, row_num):
            if not class_name or not department:
                return None
            map_key = f'{department.name}|{class_name}|{grade_text}'
            mapped_name = str(class_map.get(map_key, class_name)).strip()
            qs = Class.objects.filter(name=mapped_name, department=department)
            class_obj = qs.filter(grade=grade_text).first() if grade_text else qs.first()
            if class_obj:
                if grade_text:
                    self._ensure_major_grade(major_obj, grade_text)
                return class_obj
            if grade_text and qs.exists():
                add_unknown('grades', f'{department.name}|{mapped_name}|{grade_text}', {
                    'row_num': row_num, 'department_name': department.name,
                    'class_name': mapped_name, 'grade': grade_text,
                })
            add_unknown('classes', f'{department.name}|{mapped_name}|{grade_text}', {
                'row_num': row_num, 'department_name': department.name,
                'class_name': mapped_name, 'grade': grade_text,
                'major_name': major_obj.name if major_obj else '',
            })
            if dry_run or not auto_create_missing:
                return None
            if not self._can_create_class(scope, department.id):
                errors.append({'row': row_num, 'sheet': '', 'message': f'无权限在院系[{department.name}]创建班级: {mapped_name}'})
                return None
            class_obj = Class.objects.create(
                name=mapped_name, department=department, major=major_obj, grade=grade_text or '',
            )
            if grade_text:
                self._ensure_major_grade(major_obj, grade_text)
            created_entities['classes'].append({
                'id': class_obj.id, 'name': class_obj.name,
                'department': department.name, 'grade': class_obj.grade,
            })
            return class_obj

        def _expand_majors_to_classes(department, major_names, row):
            """@private 将专业名称列表展开为该院系该专业下所有班级 ID。"""
            class_ids = []
            for mname in major_names:
                major = Major.objects.filter(name=mname.strip(), department=department).first()
                if not major:
                    _warn(row, f'负责专业不存在: {mname}（院系: {department.name}）')
                    continue
                cids = list(Class.objects.filter(major=major, department=department).values_list('id', flat=True))
                if not cids:
                    _warn(row, f'专业[{mname}]下暂无班级')
                class_ids.extend(cid for cid in cids if cid not in class_ids)
            return class_ids

        # ── 主处理循环（行已按 student→assistant→counselor→director 排序） ──
        for row in rows:
            row_num = row.get('row_num')
            role_code = row.get('role_code', '').strip()
            source_sheet = row.get('source_sheet', '')

            if not role_code:
                _err(row, '角色代码不能为空')
                continue
            if role_code == 'superadmin':
                _err(row, f'不允许通过导入创建{get_role_display_name(ROLE_LEVEL_SUPERADMIN)}')
                continue

            role = Role.objects.filter(code=role_code).first()
            if not role:
                _err(row, f'角色代码不存在: {role_code}')
                continue
            if role.code in dependency_graph['nodes']:
                dependency_graph['nodes'][role.code] += 1
            if scope['allowed_role_codes'] is not None and role.code not in scope['allowed_role_codes']:
                _err(row, f'当前账号无权导入该角色: {role_code}')
                continue
            if role.level >= scope['operator_level']:
                _err(row, f'无权导入等级不低于自身的角色: {role_code}（等级{role.level}）')
                continue

            # ────────────────────────────────────────────
            # 学生助理：从学生中提拔（不创建新用户）
            # ────────────────────────────────────────────
            if role_code == 'student_assistant':
                student_no = row.get('student_no', '').strip()
                if not student_no:
                    _err(row, f'{get_role_display_name(ROLE_LEVEL_ASSISTANT)}必须填写学号')
                    continue

                # 查找学生：优先本文件同批导入的，再查数据库
                cached = imported_students.get(student_no)
                if cached is True:
                    # dry_run 模式下同文件学生仅有占位标记，从数据库尝试查找
                    target_user = User.objects.filter(student_no=student_no).first()
                elif cached:
                    target_user = cached
                else:
                    target_user = User.objects.filter(student_no=student_no).first()

                if not target_user and cached is not True:
                    if student_no not in imported_students:
                        _err(row, f'学号不存在: {student_no}（请确认学生已在「学生导入」表中或已存在于系统）')
                        continue

                # 默认负责自身所在班级（跨班分配由评审规则控制，导入时不指定）
                responsible_class_ids = []
                if target_user and target_user.class_obj_id:
                    responsible_class_ids = [target_user.class_obj_id]

                dependency_graph['edges']['assistant_to_class'] += len(responsible_class_ids)
                if len(dependency_graph['samples']) < 20:
                    dependency_graph['samples'].append(
                        f'{_row_label(row)} 学生助理[{student_no}] 负责班级数: {len(responsible_class_ids)}'
                    )

                if dry_run:
                    success_count += 1
                    processed_count += 1
                    continue

                try:
                    student_role = Role.objects.filter(code='student').first()
                    role_ids = [r for r in [student_role.id if student_role else None, role.id] if r]
                    sync_serializer._sync_user_roles(target_user, role_ids, responsible_class_ids)
                    success_count += 1
                except Exception as exc:
                    _err(row, f'提拔{get_role_display_name(ROLE_LEVEL_ASSISTANT)}失败: {str(exc)}')
                processed_count += 1
                if progress_callback and not dry_run:
                    progress_callback(processed_count, total_rows, success_count)
                continue

            # ────────────────────────────────────────────
            # 学生 / 评审老师 / 院系主任：创建新用户
            # ────────────────────────────────────────────
            username = row.get('username', '').strip()
            student_no = row.get('student_no', '').strip()
            employee_no = row.get('employee_no', '').strip()
            full_name = row.get('full_name', '').strip()
            raw_gender = row.get('gender', '').strip()
            email = row.get('email', '').strip()
            phone = row.get('phone', '').strip()
            password = row.get('password', '').strip() or '123456'

            if not username:
                _err(row, '用户名不能为空')
                continue
            if username in seen_usernames:
                _err(row, f'导入文件内用户名重复: {username}')
                continue
            seen_usernames.add(username)
            if User.objects.filter(username=username).exists():
                _err(row, f'用户名已存在: {username}')
                continue

            if role_code == 'student':
                if not student_no:
                    _err(row, '学生必须填写学号')
                    continue
            if role_code in ('counselor', 'director'):
                if not employee_no:
                    _err(row, f'{get_role_display_name(ROLE_LEVEL_COUNSELOR)}/{get_role_display_name(ROLE_LEVEL_DIRECTOR)}必须填写工号')
                    continue

            if student_no:
                if student_no in seen_student_no:
                    _err(row, f'导入文件内学号重复: {student_no}')
                    continue
                seen_student_no.add(student_no)
                if User.objects.filter(student_no=student_no).exists():
                    _err(row, f'学号已存在: {student_no}')
                    continue
            if employee_no:
                if employee_no in seen_employee_no:
                    _err(row, f'导入文件内工号重复: {employee_no}')
                    continue
                seen_employee_no.add(employee_no)
                if User.objects.filter(employee_no=employee_no).exists():
                    _err(row, f'工号已存在: {employee_no}')
                    continue

            if raw_gender not in gender_map:
                _err(row, f'性别值无效: {raw_gender}（可填 男/女/未知）')
                continue

            # ── 院系主任特殊处理：按负责院系自动汇总 ──
            if role_code == 'director':
                dept_names = self._split_multi_values(row.get('responsible_departments'))
                if not dept_names:
                    _err(row, f'{get_role_display_name(ROLE_LEVEL_DIRECTOR)}必须填写负责院系名称')
                    continue

                departments = []
                for dname in dept_names:
                    dept = resolve_department(dname, row_num)
                    if dept:
                        departments.append(dept)
                    elif not scope['is_super']:
                        _err(row, f'院系不存在: {dname}')

                if not departments:
                    if scope['is_super']:
                        # 超管模式：院系尚未创建，已加入 unknown，不阻断
                        _warn(row, f'负责院系尚未创建，请在下方创建院系后重新预检')
                    else:
                        if not errors or errors[-1].get('row') != row_num:
                            _err(row, '未找到有效的负责院系')
                    continue

                primary_dept = departments[0]
                dept_ids = [d.id for d in departments]

                # 院系主任直接管辖负责院系下的所有班级
                all_class_ids = list(
                    Class.objects.filter(department_id__in=dept_ids)
                    .values_list('id', flat=True)
                )

                if not all_class_ids:
                    dept_name_str = ', '.join(d.name for d in departments)
                    _warn(row, f'院系[{dept_name_str}]下暂无班级，主任暂无管辖班级')

                dependency_graph['edges']['director_to_dept'] += len(departments)
                if len(dependency_graph['samples']) < 20:
                    dependency_graph['samples'].append(
                        f'{_row_label(row)} 主任[{username}] 负责院系: {", ".join(d.name for d in departments)}, 汇总班级数: {len(all_class_ids)}'
                    )

                if dry_run:
                    success_count += 1
                    processed_count += 1
                    continue

                try:
                    new_user = User(
                        username=username, email=email or '',
                        name=full_name, first_name='', last_name='', phone=phone,
                        student_no=None, employee_no=employee_no or None,
                        gender=gender_map.get(raw_gender, ''),
                        department=primary_dept, class_obj=None,
                        is_active=scope['is_super'],
                        must_change_password=force_change_password,
                    )
                    new_user.password = _hash_password(password)
                    new_user.save()
                    sync_serializer._sync_user_roles(new_user, [role.id], all_class_ids)
                    for dept_obj in departments:
                        UserRole.objects.get_or_create(
                            user=new_user, role=role,
                            scope_id=dept_obj.id, scope_type='department',
                        )
                    success_count += 1
                except Exception as exc:
                    _err(row, f'创建用户失败: {str(exc)}')
                processed_count += 1
                if progress_callback and not dry_run:
                    progress_callback(processed_count, total_rows, success_count)
                continue

            # ── 学生 / 评审老师通用处理 ──
            dept_name = row.get('department_name', '').strip()
            class_name = row.get('class_name', '').strip()
            major_name = row.get('major_name', '').strip()
            grade_text = row.get('grade', '').strip()

            department = resolve_department(dept_name, row_num)
            if not department and not scope['is_super']:
                _err(row, '院系名称不能为空或不存在')
                continue
            if department and scope['allowed_dept_ids'] is not None and department.id not in scope['allowed_dept_ids']:
                _err(row, f'您无权向该院系（{department.name}）导入用户')
                continue

            major_obj = resolve_major(department, major_name, row_num) if major_name else None
            class_obj = resolve_class(department, major_obj, class_name, grade_text, row_num) if class_name else None
            if class_obj and scope['allowed_class_ids'] is not None and class_obj.id not in scope['allowed_class_ids']:
                _err(row, f'您无权向该班级（{class_obj.name}）导入用户')
                continue

            responsible_class_ids = []

            if role_code == 'student':
                if scope['is_counselor'] and not class_obj:
                    _err(row, f'{get_role_display_name(ROLE_LEVEL_COUNSELOR)}导入学生时必须指定有效班级')
                    continue
                if class_obj:
                    dependency_graph['edges']['student_to_class'] += 1
                    if len(dependency_graph['samples']) < 20:
                        dependency_graph['samples'].append(f'{_row_label(row)} 学生[{username}] -> 班级[{class_obj.name}]')

            elif role_code == 'counselor':
                resp_major_names = self._split_multi_values(row.get('responsible_majors'))
                managed_class_names = self._split_multi_values(row.get('responsible_classes'))

                if not department and (resp_major_names or managed_class_names):
                    for cname in managed_class_names:
                        add_unknown('classes', f'(待创建院系){dept_name}|{cname}|', {
                            'row_num': row_num, 'department_name': dept_name,
                            'class_name': cname, 'grade': '',
                            'major_name': '',
                        })
                    for mname in resp_major_names:
                        add_unknown('majors', f'{dept_name}|{mname}', {
                            'row_num': row_num, 'department_name': dept_name,
                            'major_name': mname,
                        })
                    _warn(row, f'院系[{dept_name}]尚未创建，负责专业/班级待院系创建后方可解析；请在下方创建院系后重新预检')
                elif managed_class_names:
                    # 显式指定了班级 → 仅使用这些班级，不展开专业列表
                    for cname in managed_class_names:
                        cls = resolve_class(department, major_obj, cname, '', row_num)
                        if cls and cls.id not in responsible_class_ids:
                            responsible_class_ids.append(cls.id)
                        elif not cls:
                            _err(row, f'负责班级不存在: {cname}')
                elif resp_major_names and department:
                    # 未指定班级、仅填了专业 → 展开为该专业下所有班级
                    responsible_class_ids = _expand_majors_to_classes(department, resp_major_names, row)
                else:
                    _err(row, f'{get_role_display_name(ROLE_LEVEL_COUNSELOR)}需通过「负责专业列表」或「负责班级列表」配置至少一个负责班级')
                    continue

                if not responsible_class_ids and (managed_class_names or resp_major_names):
                    _warn(row, '未能解析出有效的负责班级')


                if scope['allowed_class_ids'] is not None:
                    out_of_scope = [cid for cid in responsible_class_ids if cid not in scope['allowed_class_ids']]
                    if out_of_scope:
                        _err(row, f'负责班级超出当前账号管辖范围')
                        continue

                dependency_graph['edges']['counselor_to_class'] += len(responsible_class_ids)
                if len(dependency_graph['samples']) < 20:
                    dependency_graph['samples'].append(
                        f'{_row_label(row)} 评审老师[{username}] 负责班级数: {len(responsible_class_ids)}'
                    )

            if dry_run:
                success_count += 1
                processed_count += 1
                if role_code == 'student' and student_no:
                    imported_students[student_no] = True
                continue

            try:
                new_user = User(
                    username=username, email=email or '',
                    name=full_name, first_name='', last_name='', phone=phone,
                    student_no=student_no or None, employee_no=employee_no or None,
                    gender=gender_map.get(raw_gender, ''),
                    department=department, class_obj=class_obj,
                    is_active=scope['is_super'],
                    must_change_password=force_change_password,
                )
                new_user.password = _hash_password(password)
                new_user.save()
                sync_serializer._sync_user_roles(new_user, [role.id], responsible_class_ids)
                success_count += 1

                if role_code == 'student' and student_no:
                    imported_students[student_no] = new_user
                if role_code == 'counselor' and department:
                    imported_counselors_by_dept.setdefault(department.id, []).append(new_user)
            except Exception as exc:
                _err(row, f'创建用户失败: {str(exc)}')

            processed_count += 1
            if progress_callback and not dry_run:
                progress_callback(processed_count, total_rows, success_count)

        unknown_entities = {
            'departments': [
                {'name': v['name'], 'rows': sorted(list(v['rows']))}
                for v in unknown['departments'].values()
            ],
            'majors': [
                {'department_name': v['department_name'], 'major_name': v['major_name'], 'rows': sorted(list(v['rows']))}
                for v in unknown['majors'].values()
            ],
            'classes': [
                {
                    'department_name': v['department_name'],
                    'major_name': v.get('major_name', ''),
                    'class_name': v['class_name'],
                    'grade': v['grade'],
                    'rows': sorted(list(v['rows'])),
                }
                for v in unknown['classes'].values()
            ],
            'grades': [
                {'department_name': v['department_name'], 'class_name': v['class_name'], 'grade': v['grade'], 'rows': sorted(list(v['rows']))}
                for v in unknown['grades'].values()
            ],
        }

        return {
            'row_count': len(rows),
            'success_count': success_count,
            'errors': errors,
            'warnings': warnings,
            'unknown_entities': unknown_entities,
            'dependency_graph': dependency_graph,
            'created_entities': created_entities,
        }

    def post(self, request):
        """用户导入：支持 dry_run 预检会话、映射修复与按权限创建缺失组织项。"""
        scope, err_resp = self._operator_scope(request)
        if err_resp:
            return err_resp

        dry_run = self._is_truthy(request.data.get('dry_run', ''))
        preview_token = str(request.data.get('preview_token') or '').strip()
        auto_create_missing = self._is_truthy(request.data.get('auto_create_missing', ''))
        resolution_payload = self._parse_resolution_payload(request)

        # 解析 excluded_rows（格式: JSON 数组, 元素为 "sheet_name:row_num" 字符串）
        excluded_rows_raw = request.data.get('excluded_rows', '')
        excluded_set = set()
        if excluded_rows_raw:
            try:
                excluded_list = json.loads(excluded_rows_raw) if isinstance(excluded_rows_raw, str) else excluded_rows_raw
                excluded_set = set(str(x) for x in excluded_list)
            except (json.JSONDecodeError, TypeError):
                pass

        file_obj = request.FILES.get('file')
        source_rows = None
        source_file_name = file_obj.name if file_obj else ''

        if preview_token and not file_obj:
            cache_key = self._preview_cache_key(request.user.id, preview_token)
            snapshot = cache.get(cache_key)
            if not snapshot:
                return Response({'detail': '预检会话不存在或已过期，请重新预检'}, status=status.HTTP_400_BAD_REQUEST)
            if snapshot.get('user_id') != request.user.id:
                return Response({'detail': '预检会话不属于当前账号'}, status=status.HTTP_403_FORBIDDEN)
            source_rows = snapshot.get('rows', [])
            source_file_name = snapshot.get('file_name', 'from_preview.xlsx')
        elif file_obj:
            source_rows = self._load_rows_from_excel(file_obj)
        else:
            return Response({'detail': '请提供 file 或 preview_token'}, status=status.HTTP_400_BAD_REQUEST)

        # 根据 excluded_rows 过滤行
        if excluded_set and source_rows:
            source_rows = [
                r for r in source_rows
                if f"{r.get('source_sheet', '')}:{r.get('row_num', '')}" not in excluded_set
            ]

        # 预检模式：仅校验，不写库，并创建会话 token
        if dry_run:
            result = self._run_import(
                request=request,
                rows=source_rows,
                scope=scope,
                dry_run=True,
                resolution=resolution_payload,
                auto_create_missing=False,
            )
            token = str(uuid.uuid4())
            cache.set(
                self._preview_cache_key(request.user.id, token),
                {
                    'user_id': request.user.id,
                    'rows': source_rows,
                    'file_name': source_file_name,
                    'created_at': timezone.now().isoformat(),
                },
                timeout=_IMPORT_PREVIEW_TTL,
            )

            log_action(
                user=request.user,
                action='users_import_preview',
                module=OperationLog.MODULE_USERS,
                level=OperationLog.LEVEL_NOTICE,
                target_type='import_preview',
                target_id=0,
                target_repr=source_file_name or 'preview',
                extra={
                    'row_count': result['row_count'],
                    'error_count': len(result['errors']),
                    'unknown_count': (
                        len(result['unknown_entities']['departments'])
                        + len(result['unknown_entities']['majors'])
                        + len(result['unknown_entities']['classes'])
                        + len(result['unknown_entities']['grades'])
                    ),
                    'preview_token': token,
                },
                request=request,
            )
            return Response({
                'status': 'preview',
                'file_name': source_file_name,
                'preview_token': token,
                'row_count': result['row_count'],
                'success_count': result['success_count'],
                'error_count': len(result['errors']),
                'warning_count': len(result['warnings']),
                'error_log': result['errors'],
                'warning_log': result['warnings'],
                'unknown_entities': result['unknown_entities'],
                'dependency_graph': result['dependency_graph'],
                'detail': '预检完成，未写入数据库',
            }, status=status.HTTP_200_OK)

        # 正式导入：后台线程异步执行
        import threading
        import time as _time

        raw_iterations = request.data.get('hash_iterations', '')
        hash_iterations = None
        if raw_iterations not in (None, '', 'default', '0'):
            try:
                val = int(raw_iterations)
                if val > 0:
                    hash_iterations = val
            except (ValueError, TypeError):
                pass

        force_change_pwd = request.data.get('force_change_password', '1') not in ('0', 'false', 'False')

        batch = ImportedUserBatch.objects.create(
            uploaded_by=request.user,
            file_name=source_file_name or 'from_preview.xlsx',
            status='processing',
            row_count=len(source_rows),
            hash_iterations=hash_iterations,
        )

        operator_user_id = request.user.id
        operator_is_super = scope['is_super']

        def _bg_import():
            """后台线程：执行导入并通过 SSE 推送进度。"""
            import django.db
            from realtime.registry import publish

            _last_push = _time.monotonic()
            _PUSH_INTERVAL = 1.5  # SSE 推送最短间隔（秒）

            def _progress_cb(current, total, ok_count):
                """进度回调：限频推送 SSE + 更新 DB。"""
                nonlocal _last_push
                now = _time.monotonic()
                is_last = (current >= total)
                if not is_last and (now - _last_push) < _PUSH_INTERVAL and current % 50 != 0:
                    return
                _last_push = now

                ImportedUserBatch.objects.filter(id=batch.id).update(
                    current_count=current, success_count=ok_count,
                )
                publish(operator_user_id, {
                    'type': 'import_progress',
                    'batch_id': batch.id,
                    'current': current,
                    'total': total,
                    'success_count': ok_count,
                    'status': 'processing',
                })

            try:
                result = self._run_import(
                    request=request,
                    rows=source_rows,
                    scope=scope,
                    dry_run=False,
                    resolution=resolution_payload,
                    auto_create_missing=auto_create_missing,
                    hash_iterations=hash_iterations,
                    force_change_password=force_change_pwd,
                    progress_callback=_progress_cb,
                )
                batch.status = 'completed'
                batch.current_count = len(source_rows)
                batch.success_count = result['success_count']
                batch.error_log = result['errors']
                batch.warning_log = result['warnings']
                batch.save(update_fields=[
                    'status', 'current_count', 'success_count', 'error_log', 'warning_log',
                ])

                _op_user = User.objects.get(id=operator_user_id)
                log_action(
                    user=_op_user,
                    action='users_import_commit',
                    module=OperationLog.MODULE_USERS,
                    level=OperationLog.LEVEL_WARNING,
                    target_type='imported_user_batch',
                    target_id=batch.id,
                    target_repr=source_file_name,
                    extra={
                        'row_count': batch.row_count,
                        'success_count': result['success_count'],
                        'error_count': len(result['errors']),
                        'warning_count': len(result['warnings']),
                        'operator_level': scope['operator_level'],
                        'pending_activation': not operator_is_super and result['success_count'] > 0,
                        'auto_create_missing': auto_create_missing,
                        'created_entities': result['created_entities'],
                        'preview_token': preview_token or '',
                        'hash_iterations': hash_iterations,
                    },
                    is_audit_event=True,
                    request=request,
                )
                if auto_create_missing and any(result['created_entities'].values()):
                    log_action(
                        user=_op_user,
                        action='users_import_org_create',
                        module=OperationLog.MODULE_USERS,
                        level=OperationLog.LEVEL_WARNING,
                        target_type='org',
                        target_id=0,
                        target_repr='import-auto-create',
                        extra=result['created_entities'],
                        is_audit_event=True,
                        request=request,
                    )

                publish(operator_user_id, {
                    'type': 'import_progress',
                    'batch_id': batch.id,
                    'current': len(source_rows),
                    'total': len(source_rows),
                    'success_count': result['success_count'],
                    'error_count': len(result['errors']),
                    'warning_count': len(result['warnings']),
                    'status': 'completed',
                    'notice': (
                        f'已成功导入 {result["success_count"]} 名用户，账号默认未激活，请联系{get_role_display_name(ROLE_LEVEL_SUPERADMIN)}审批激活后方可登录。'
                        if not operator_is_super and result['success_count'] > 0
                        else ''
                    ),
                })
            except Exception as exc:
                batch.status = 'failed'
                batch.error_log = [{'message': f'文件处理失败: {str(exc)}'}]
                batch.save(update_fields=['status', 'error_log'])
                publish(operator_user_id, {
                    'type': 'import_progress',
                    'batch_id': batch.id,
                    'current': 0,
                    'total': len(source_rows),
                    'success_count': 0,
                    'status': 'failed',
                    'detail': str(exc),
                })
            finally:
                if preview_token:
                    cache.delete(self._preview_cache_key(operator_user_id, preview_token))
                django.db.connection.close()

        t = threading.Thread(target=_bg_import, daemon=True, name=f'import-batch-{batch.id}')
        t.start()

        return Response({
            'status': 'processing',
            'batch': ImportedUserBatchSerializer(batch).data,
            'detail': '导入已在后台启动，请通过进度条跟踪状态',
        }, status=status.HTTP_202_ACCEPTED)


class ImportProgressView(APIView):
    """
    GET /api/v1/users/import/progress/<batch_id>/
    轮询导入进度（SSE 断线降级方案）。
    若检测到导入线程已不存在（如服务器重启），自动将批次标记为 failed。
    """
    permission_classes = [IsAuthenticated]

    @staticmethod
    def _is_import_thread_alive(batch_id):
        """检查对应的后台导入线程是否仍在运行。"""
        import threading
        expected_name = f'import-batch-{batch_id}'
        return any(t.name == expected_name for t in threading.enumerate())

    def get(self, request, batch_id):
        try:
            batch = ImportedUserBatch.objects.get(id=batch_id)
        except ImportedUserBatch.DoesNotExist:
            return Response({'detail': '批次不存在'}, status=status.HTTP_404_NOT_FOUND)
        if batch.uploaded_by_id != request.user.id:
            return Response({'detail': '无权查看该批次'}, status=status.HTTP_403_FORBIDDEN)

        if batch.status == 'processing' and not self._is_import_thread_alive(batch.id):
            batch.status = 'failed'
            batch.error_log = (batch.error_log or []) + [{
                'row': None,
                'sheet': '',
                'message': '导入因服务器重启而中断，已处理的数据已写入。请检查数据完整性后决定是否重新导入。',
            }]
            batch.save(update_fields=['status', 'error_log'])

        return Response({
            'batch_id': batch.id,
            'status': batch.status,
            'current': batch.current_count,
            'total': batch.row_count,
            'success_count': batch.success_count,
            'error_count': len(batch.error_log) if batch.error_log else 0,
            'warning_count': len(batch.warning_log) if batch.warning_log else 0,
            'error_log': batch.error_log if batch.status in ('completed', 'failed') else [],
            'warning_log': batch.warning_log if batch.status == 'completed' else [],
        })


class LoginHistoryView(APIView):
    """
    GET /api/v1/users/me/login-history/
    返回当前用户最近 10 条登录记录，数据来源为操作日志中 module=auth, action=login 的条目。
    """
    permission_classes = [IsAuthenticated]

    def get(self, request):
        logs = (
            OperationLog.objects
            .filter(user=request.user, module='auth', action='login')
            .order_by('-created_at')
            .values('id', 'ip_address', 'external_ip', 'user_agent', 'created_at')[:10]
        )
        return Response(list(logs))


class UpdateProfileView(APIView):
    """
    PATCH /api/v1/users/me/profile/
    允许当前用户修改自己的基本信息（仅限 email、phone）。
    用户名、学号、工号、院系、班级、角色等管理员字段不允许自行修改。
    """
    permission_classes = [IsAuthenticated]

    def patch(self, request):
        user = request.user
        email = request.data.get('email')
        phone = request.data.get('phone')
        changed = {}

        if email is not None:
            email = email.strip()
            if email and '@' not in email:
                return Response(
                    {'detail': '邮箱格式不正确'},
                    status=status.HTTP_400_BAD_REQUEST,
                )
            changed['email'] = {'old': user.email, 'new': email}
            user.email = email

        if phone is not None:
            phone = phone.strip()
            if phone and (not phone.isdigit() or len(phone) < 7 or len(phone) > 20):
                return Response(
                    {'detail': '手机号格式不正确'},
                    status=status.HTTP_400_BAD_REQUEST,
                )
            changed['phone'] = {'old': user.phone, 'new': phone}
            user.phone = phone

        if not changed:
            return Response(
                {'detail': '未提供需要更新的字段（可更新: email, phone）'},
                status=status.HTTP_400_BAD_REQUEST,
            )

        update_fields = list(changed.keys())
        user.save(update_fields=update_fields)

        log_action(
            user=user,
            action='profile_update',
            module=OperationLog.MODULE_USERS,
            level=OperationLog.LEVEL_INFO,
            target_type='user',
            target_id=user.id,
            target_repr=user.username,
            extra={'changed': changed},
            request=request,
        )
        return Response(UserSerializer(user).data)
