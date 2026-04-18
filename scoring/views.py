"""
审核与评分 API：初审、双评、仲裁、导入。
"""
import json
import os
from decimal import Decimal

from django.db import transaction
from django.db.models import Prefetch, Q
from rest_framework import status
from rest_framework.views import APIView
from rest_framework.generics import ListAPIView, RetrieveAPIView
from rest_framework.permissions import IsAuthenticated
from rest_framework.response import Response
from rest_framework.parsers import MultiPartParser, FormParser

from audit.models import OperationLog, ImportPermissionRequest
from audit.services import log_action
from submission.models import StudentSubmission, SubmissionAnswer, Evidence
from submission.serializers import StudentSubmissionSerializer
from eval.models import EvalProject, EvalIndicator, ReviewRule
from .models import (
    ScoreRecord,
    ArbitrationRecord,
    ImportedScoreBatch,
    ImportedScoreDetail,
    ReviewAssignment,
    ReviewObjection,
    ReviewObjectionAttachment,
)
from .serializers import (
    ScoreRecordSerializer,
    ImportedScoreBatchSerializer,
    ReviewQuestionSerializer,
    ReviewAssignmentSerializer,
    ReviewTaskSubmissionSerializer,
    ReviewObjectionSerializer,
    ReviewObjectionAttachmentSerializer,
)
from .services import recompute_submission_final_score, submission_missing_required_leaf_indicators
from eval.utils import raw_max_score as _raw_max_score
from users.permissions import user_level_at_least, user_is_admin, user_is_super_admin
from users.role_resolver import (
    ROLE_LEVEL_ASSISTANT, ROLE_LEVEL_COUNSELOR, ROLE_LEVEL_DIRECTOR, ROLE_LEVEL_SUPERADMIN,
    get_role_display_name,
)
from .assignment_services import (
    generate_review_assignments_for_project,
    _resolve_candidates_for_submission,
    _pick_assistants_for_submission,
)


def _latest_project_assignment_version(project_id):
    """
    @deprecated 版本号过滤已废弃，保留函数仅供历史兼容参考。
    """
    return (
        ReviewAssignment.objects.filter(project_id=project_id)
        .order_by('-assignment_version')
        .values_list('assignment_version', flat=True)
        .first()
    )


def _assigned_submission_ids_for_user(user, project_id=None):
    """
    返回用户当前 status='assigned' 的 submission_id 集合。
    """
    qs = ReviewAssignment.objects.filter(reviewer=user, status='assigned')
    if project_id is not None:
        qs = qs.filter(project_id=project_id)
    return set(qs.values_list('submission_id', flat=True).distinct())


def _get_active_assignment_for_submission(user, submission):
    """
    返回用户在该 submission 上当前有效的分配任务（status='assigned'），若无则返回 None。
    """
    return (
        ReviewAssignment.objects.filter(
            reviewer=user,
            submission=submission,
            project_id=submission.project_id,
            status='assigned',
        )
        .order_by('round_type', 'id')
        .first()
    )


def _get_readable_assignment_for_submission(user, submission):
    """
    返回用户在该 submission 上可读的任务（assigned/completed）。
    用于评分完成后的历史查看权限判定。
    """
    return (
        ReviewAssignment.objects.filter(
            reviewer=user,
            submission=submission,
            project_id=submission.project_id,
            status__in=['assigned', 'completed'],
        )
        .order_by('status', 'round_type', 'id')
        .first()
    )


def _project_has_assignments(project_id):
    return ReviewAssignment.objects.filter(project_id=project_id).exists()


def _user_max_role_level(user):
    """
    返回用户所有角色中的最高等级（用于判断降级场景）。
    @param user {User}
    @returns {int}
    """
    from users.models import UserRole
    from django.db.models import Max
    result = UserRole.objects.filter(user=user).select_related('role').aggregate(
        max_level=Max('role__level')
    )
    return result['max_level'] or -1


def _build_score_context(user):
    """
    构建评分上下文字段，用于写入 ScoreRecord。
    @param user {User}
    @returns {dict} 包含 scorer_role_level, scorer_max_role_level, is_delegated
    """
    current_level = user.current_role.level if user.current_role else -1
    max_level = _user_max_role_level(user)
    return {
        'scorer_role_level': current_level if current_level >= 0 else None,
        'scorer_max_role_level': max_level if max_level >= 0 else None,
        'is_delegated': max_level > current_level,
    }


def _is_single_review(rule):
    """判断是否单评模式。"""
    return bool(rule) and not bool(rule.dual_review_enabled)


def _single_allowed_role_type(rule):
    """返回单评模式允许的常规任务角色。"""
    mode = (getattr(rule, 'single_review_mode', None) or 'assistant_single') if rule else 'assistant_single'
    return 'counselor' if mode == 'counselor_single' else 'assistant'


def _maybe_create_single_review_confirm(submission, reviewer, assignment):
    """
    单评模式下，仅在"异常"情况（如提交人本身是助理角色，即助理自评）时
    自动为辅导员生成 counselor_confirm 任务。
    正常提交不强制确认，辅导员可通过仲裁接口主动介入（即"异常可上报"）。
    """
    if assignment is None or assignment.role_type != 'assistant':
        return
    rule = getattr(submission.project, 'review_rule', None)
    if not rule or rule.dual_review_enabled:
        return

    from users.models import UserRole as _CheckUR
    is_assistant_sub = _CheckUR.objects.filter(
        user_id=submission.user_id,
        role__level=ROLE_LEVEL_ASSISTANT,
    ).exists()
    if not is_assistant_sub:
        return

    leaf_indicator_ids = _self_leaf_indicator_ids(submission.project, role_type=assignment.role_type)
    if not leaf_indicator_ids:
        return
    scored_count = ScoreRecord.objects.filter(
        submission=submission,
        reviewer=reviewer,
        indicator_id__in=leaf_indicator_ids,
        score_channel='assignment',
    ).exclude(round_type=3).values_list('indicator_id', flat=True).distinct().count()

    if scored_count < len(leaf_indicator_ids):
        return

    from .assignment_services import _find_counselor_owner
    class_id = getattr(submission.user, 'class_obj_id', None)
    if not class_id:
        return
    counselor = _find_counselor_owner(class_id)
    if counselor is None:
        return
    ReviewAssignment.objects.get_or_create(
        submission=submission,
        project=submission.project,
        reviewer=counselor,
        role_type='counselor_confirm',
        round_type=4,
        defaults={
            'strategy_mode': 'single_review_confirm',
            'assignment_version': 1,
            'assigned_by': reviewer,
            'status': 'assigned',
        },
    )


def _self_leaf_indicator_ids(project, role_type=None):
    """获取项目下需要评审打分的叶子指标 ID（支持仅记录模块可配置进入评审）。"""
    qs = EvalIndicator.objects.filter(project=project, score_source='self').filter(
        Q(is_record_only=False) | Q(is_record_only=True, record_only_requires_review=True)
    )
    return list(
        qs
        .exclude(
            id__in=EvalIndicator.objects.filter(
                project=project, parent__isnull=False
            ).values_list('parent_id', flat=True)
        )
        .values_list('id', flat=True)
    )


def _indicator_requires_review(indicator):
    """判断指标是否需要进入评审链路。"""
    if not indicator.is_record_only:
        return True
    return bool(getattr(indicator, 'record_only_requires_review', False))


def _maybe_mark_assignment_completed(submission, reviewer, assignment):
    """当当前评审人已完成全部叶子指标评分时，自动完成该评审任务。"""
    if assignment is None or assignment.status != 'assigned':
        return
    if assignment.role_type not in {'assistant', 'counselor', 'counselor_confirm'}:
        return
    indicator_ids = _self_leaf_indicator_ids(submission.project, role_type=assignment.role_type)
    if not indicator_ids:
        return
    scored_count = ScoreRecord.objects.filter(
        submission=submission,
        reviewer=reviewer,
        round_type=assignment.round_type,
        score_channel='assignment',
        indicator_id__in=indicator_ids,
    ).exclude(round_type=3).values_list('indicator_id', flat=True).distinct().count()
    if scored_count < len(indicator_ids):
        return
    from django.utils import timezone
    assignment.status = 'completed'
    assignment.completed_at = timezone.now()
    assignment.save(update_fields=['status', 'completed_at', 'updated_at'])


def _maybe_finalize_submission_status(submission):
    """当常规评审任务均完成且已有最终分时，将提交状态闭环到 approved。"""
    submission.refresh_from_db(fields=['status', 'final_score'])
    if submission.final_score is None:
        return
    if submission.status not in ('submitted', 'under_review'):
        return
    has_pending_assignment = ReviewAssignment.objects.filter(
        submission=submission,
        status='assigned',
        role_type__in=['assistant', 'counselor', 'counselor_confirm'],
    ).exists()
    if has_pending_assignment:
        return
    missing_required = submission_missing_required_leaf_indicators(submission)
    if missing_required:
        return
    submission.status = 'approved'
    submission.save(update_fields=['status'])


def _single_mode_conflict_exists(submission, indicator, reviewer_id, round_type):
    """单评模式下检查同指标是否已存在其他常规评分。"""
    return ScoreRecord.objects.filter(
        submission=submission,
        indicator=indicator,
    ).exclude(
        round_type=3,
    ).exclude(
        reviewer_id=reviewer_id,
        round_type=round_type,
    ).exists()


def _to_decimal_or_none(value):
    if value in (None, ''):
        return None
    try:
        return Decimal(str(value))
    except Exception:
        return None


def _validate_indicator_score(indicator, raw_score):
    """
    统一校验题目分值合法区间。
    - 下限固定为 0
    - 上限取 raw_max_score（None 表示无上限）
    @returns {(Decimal|None, str|None)} (标准化分数, 错误消息)
    """
    score = _to_decimal_or_none(raw_score)
    if score is None:
        return None, f'指标「{indicator.name}」分数格式无效，请输入数字'

    min_score = Decimal('0')
    max_score = _to_decimal_or_none(_raw_max_score(indicator))
    if score < min_score:
        return None, f'指标「{indicator.name}」分数越界，合法区间为 [{min_score}, {"无限制" if max_score is None else max_score}]'
    if max_score is not None and score > max_score:
        return None, f'指标「{indicator.name}」分数越界，合法区间为 [{min_score}, {max_score}]'
    return score, None


def _format_logical_round_label(logical_round_type):
    """
    将逻辑轮次值格式化为展示文案。
    @param logical_round_type {int|str|None}
    @returns {str}
    """
    if logical_round_type == 3:
        return '仲裁'
    if logical_round_type == 'import':
        return '导入'
    if isinstance(logical_round_type, int) and logical_round_type > 0:
        return f'{logical_round_type}评'
    return '未知轮次'


def _attach_logical_round_fields(score_rows):
    """
    为评分记录补充逻辑轮次字段。
    规则：
    - 常规评审（score_channel=assignment 且非仲裁）按 submission+reviewer 的首次评分时间排序映射为 1评/2评...
    - 仲裁固定为 round=3，导入固定为 round='import'，两者不参与常规轮次映射
    @param score_rows {list[dict]}
    @returns {list[dict]}
    """
    if not score_rows:
        return score_rows

    first_seen = {}
    for idx, row in enumerate(score_rows):
        score_channel = row.get('score_channel')
        round_type = row.get('round_type')
        reviewer_id = row.get('reviewer')
        if score_channel != 'assignment' or round_type == 3 or reviewer_id is None:
            continue
        key = (row.get('submission'), reviewer_id)
        created_at = row.get('created_at') or ''
        row_id = row.get('id')
        id_key = row_id if isinstance(row_id, int) else idx
        marker = (created_at, id_key)
        prev = first_seen.get(key)
        if prev is None or marker < prev:
            first_seen[key] = marker

    order_map = {}
    for (submission_id, reviewer_id), marker in sorted(
        first_seen.items(),
        key=lambda item: (item[0][0], item[1][0], item[1][1], item[0][1]),
    ):
        reviewer_order = order_map.setdefault(submission_id, {})
        if reviewer_id not in reviewer_order:
            reviewer_order[reviewer_id] = len(reviewer_order) + 1

    for row in score_rows:
        score_channel = row.get('score_channel')
        round_type = row.get('round_type')
        reviewer_id = row.get('reviewer')
        submission_id = row.get('submission')
        if round_type == 3 or score_channel == 'arbitration':
            logical_round_type = 3
        elif score_channel == 'import':
            logical_round_type = 'import'
        else:
            logical_round_type = (
                order_map.get(submission_id, {}).get(reviewer_id)
                or (round_type if isinstance(round_type, int) and round_type > 0 else None)
            )
        row['logical_round_type'] = logical_round_type
        row['logical_round_label'] = _format_logical_round_label(logical_round_type)
    return score_rows


def _indicator_module_key(indicator):
    node = indicator
    while node is not None:
        if getattr(node, 'category', None):
            return str(node.category).strip()
        node = getattr(node, 'parent', None)
    return ''


def _overall_regular_score_diff(submission):
    """仅采集正式分配任务评分通道的记录计算总分差。"""
    from decimal import Decimal
    records = ScoreRecord.objects.filter(
        submission=submission, score_channel='assignment',
    ).exclude(round_type=3)
    totals = {}
    for rec in records:
        totals.setdefault(rec.reviewer_id, Decimal('0'))
        totals[rec.reviewer_id] += rec.score
    if len(totals) < 2:
        return None
    scores = list(totals.values())
    return max(scores) - min(scores)


def _threshold_hit_for_indicator(submission, indicator, rule):
    """仅采集正式分配任务评分通道的记录判断分差阈值。"""
    records = list(ScoreRecord.objects.filter(
        submission=submission, indicator=indicator, score_channel='assignment',
    ).exclude(round_type=3))
    if len(records) < 2:
        return False, '评分记录不足2条'

    scores = [r.score for r in records]
    indicator_diff = max(scores) - min(scores)
    module_thresholds = rule.module_diff_thresholds or {}
    module_key = _indicator_module_key(indicator)
    module_threshold = _to_decimal_or_none(module_thresholds.get(module_key)) if module_key else None
    fallback_threshold = _to_decimal_or_none(rule.score_diff_threshold)
    indicator_threshold = module_threshold if module_threshold is not None else fallback_threshold
    indicator_hit = bool(indicator_threshold is not None and indicator_diff > indicator_threshold)

    overall_threshold = _to_decimal_or_none(rule.overall_score_diff_threshold)
    overall_diff = _overall_regular_score_diff(submission) if overall_threshold is not None else None
    overall_hit = bool(overall_threshold is not None and overall_diff is not None and overall_diff > overall_threshold)

    if indicator_hit or overall_hit:
        return True, 'threshold_hit'
    if indicator_threshold is None and overall_threshold is None:
        return False, '未配置阈值'
    return False, 'threshold_not_hit'


def _user_manages_submission_class(user, submission):
    """
    判断评审老师是否管辖该提交所属班级。
    """
    from users.models import UserRole
    if not submission.user.class_obj_id:
        return False
    return UserRole.objects.filter(
        user=user,
        role__level__gte=2,
        scope_type='class',
        scope_id=submission.user.class_obj_id,
    ).exists()


def _get_counselor_class_ids(user):
    """
    获取评审老师（辅导员）负责的所有班级 id 集合。
    通过 UserRole 表 scope_type='class' 查询。
    """
    from users.models import UserRole as UserRoleModel
    return set(
        UserRoleModel.objects.filter(
            user=user, scope_type='class', scope_id__isnull=False
        ).values_list('scope_id', flat=True)
    )


def _resolve_director_department_id(user):
    """解析主任导入范围院系。"""
    if getattr(user, 'department_id', None):
        return user.department_id
    from users.models import UserRole as UserRoleModel
    director_scope = (
        UserRoleModel.objects
        .filter(user=user, role__level=ROLE_LEVEL_DIRECTOR, scope_type='department', scope_id__isnull=False)
        .order_by('id')
        .values_list('scope_id', flat=True)
        .first()
    )
    return director_scope


def _resolve_import_actor_scope(user):
    """
    解析导入权限作用域：按用户最高角色等级决定导入边界。
    """
    current_level = user.current_role.level if user.current_role else -1
    max_level = _user_max_role_level(user)
    effective_level = -1
    department_id = None
    class_ids = set()

    if max_level >= ROLE_LEVEL_SUPERADMIN:
        effective_level = ROLE_LEVEL_SUPERADMIN
    elif max_level >= ROLE_LEVEL_DIRECTOR:
        effective_level = ROLE_LEVEL_DIRECTOR
        department_id = _resolve_director_department_id(user)
    elif max_level >= ROLE_LEVEL_COUNSELOR:
        effective_level = ROLE_LEVEL_COUNSELOR
        class_ids = _get_counselor_class_ids(user)

    return {
        'current_level': current_level,
        'max_level': max_level,
        'effective_level': effective_level,
        'department_id': department_id,
        'class_ids': class_ids,
    }


def _sanitize_import_policy(config):
    """规范化导入策略配置。"""
    cfg = config or {}
    if not isinstance(cfg, dict):
        cfg = {}
    mode = str(cfg.get('import_mode', 'subordinate_self')).strip()
    if mode not in {'subordinate_self', 'upper_unified'}:
        mode = 'subordinate_self'
    requires_approval = bool(cfg.get('subordinate_requires_approval', True))
    return {
        'import_mode': mode,
        'subordinate_requires_approval': requires_approval,
    }


def _is_import_blocked_by_policy(user, project, actor_scope):
    """
    返回 (is_blocked, detail)。
    upper_unified 策略下：主任及以上可直接导入；评审老师需审批（若策略要求）。
    """
    policy = _sanitize_import_policy(getattr(project, 'import_config', {}) or {})
    if policy['import_mode'] != 'upper_unified':
        return False, ''

    if actor_scope['effective_level'] >= ROLE_LEVEL_DIRECTOR:
        return False, ''
    if not policy['subordinate_requires_approval']:
        return False, ''

    approved = ImportPermissionRequest.objects.filter(
        project=project,
        requester=user,
        status=ImportPermissionRequest.STATUS_APPROVED,
    ).exists()
    if approved:
        return False, ''
    return True, '当前项目已启用“上级统一导入”，您需向直系上级发起申请并获批后方可导入'


def _user_can_manage_project_import(user, project, actor_scope=None):
    """
    成绩导入权限判断（项目级别入口校验）：
    - 超管（level>=5）：全局可导入
    - 院系主任（level==3）：本院系有学生参与该项目即可进入，
      具体每行学生是否属于本院系在逐行处理时再判断。
    - 评审老师（level==2）：其负责班级中有学生参与该项目即可进入，
      具体每行学生是否属于其负责班级在逐行处理时再判断。
    """
    scope = actor_scope or _resolve_import_actor_scope(user)
    level = scope['effective_level']
    if level >= ROLE_LEVEL_SUPERADMIN:
        return True
    if level == ROLE_LEVEL_DIRECTOR:
        dept_id = scope['department_id']
        if not dept_id:
            return False
        return StudentSubmission.objects.filter(
            project=project,
            user__department_id=dept_id,
        ).exists()
    if level == ROLE_LEVEL_COUNSELOR:
        class_ids = scope['class_ids']
        if not class_ids:
            return False
        return StudentSubmission.objects.filter(
            project=project,
            user__class_obj_id__in=class_ids,
        ).exists()
    return False


def _get_director_for_submission(submission):
    """根据提交归属院系查找院系主任（优先院系 scope，其次班级 scope 兜底）。"""
    from users.models import UserRole
    from org.models import Class
    dept_id = getattr(submission.user, 'department_id', None)
    if not dept_id:
        return None
    ur = UserRole.objects.filter(
        scope_type='department',
        scope_id=dept_id,
        role__level=ROLE_LEVEL_DIRECTOR,
    ).select_related('user').first()
    if ur:
        return ur.user
    cls_ids = list(Class.objects.filter(department_id=dept_id).values_list('id', flat=True))
    if not cls_ids:
        return None
    ur = UserRole.objects.filter(
        scope_type='class',
        scope_id__in=cls_ids,
        role__level=ROLE_LEVEL_DIRECTOR,
    ).select_related('user').first()
    return ur.user if ur else None


def _can_user_access_objection(user, objection):
    """异议工单可见性判断。"""
    if user_is_admin(user):
        return True
    if objection.raised_by_id == user.id:
        return True
    if objection.assigned_to_id == user.id:
        return True
    if user_level_at_least(user, ROLE_LEVEL_DIRECTOR):
        sub_dept_id = getattr(objection.submission.user, 'department_id', None)
        if sub_dept_id and sub_dept_id == getattr(user, 'department_id', None):
            return True
    if user_level_at_least(user, ROLE_LEVEL_COUNSELOR):
        return _user_manages_submission_class(user, objection.submission)
    return False


def _triggered_reason_by_actor_level(level):
    """根据处理人等级映射仲裁来源。"""
    if level >= ROLE_LEVEL_SUPERADMIN:
        return 'admin_objection'
    if level >= ROLE_LEVEL_DIRECTOR:
        return 'director_objection'
    if level >= ROLE_LEVEL_COUNSELOR:
        return 'counselor_objection'
    return 'assistant_objection'


def _parse_excel_rows(file_obj):
    """
    解析 Excel 文件，返回 (header_row, data_rows)。
    header_row: list[str]，data_rows: list[tuple]。
    若文件为空或无数据行则抛出 ValueError。
    """
    import openpyxl
    wb = openpyxl.load_workbook(file_obj, read_only=True)
    ws = wb.active
    all_rows = list(ws.iter_rows(min_row=1, values_only=True))
    if not all_rows:
        raise ValueError('Excel 文件为空，请使用下载的导入模板填写后上传')
    header_row = [str(c).strip() if c is not None else '' for c in all_rows[0]]
    data_rows = all_rows[1:]
    return header_row, data_rows


class ReviewTaskListAPIView(ListAPIView):
    """GET /api/v1/review/tasks/ 待我审核/已审核列表（评审老师及以上 level>=2）。"""
    permission_classes = [IsAuthenticated]
    serializer_class = ReviewTaskSubmissionSerializer

    def get_queryset(self):
        user = self.request.user
        if not user_level_at_least(user, 2):
            return StudentSubmission.objects.none()

        qs = StudentSubmission.objects.filter(
            status__in=['submitted', 'under_review', 'approved', 'rejected']
        ).exclude(
            via_late_channel=True, status='submitted'
        ).select_related('project', 'user', 'user__class_obj', 'user__department').order_by('-updated_at')

        # 严格按 current_role.level 决定可见范围（不使用 _user_max_role_level 回退）
        if user_is_admin(user):
            pass
        elif user_level_at_least(user, ROLE_LEVEL_DIRECTOR):
            from users.models import UserRole
            from org.models import Class
            dept_ids = list(
                UserRole.objects.filter(user=user, scope_type='department')
                .values_list('scope_id', flat=True)
            )
            class_ids_from_dept = list(
                Class.objects.filter(department_id__in=dept_ids).values_list('id', flat=True)
            ) if dept_ids else []
            scope_class_ids = list(
                UserRole.objects.filter(user=user, scope_type='class')
                .values_list('scope_id', flat=True)
            )
            all_class_ids = list(set(class_ids_from_dept + scope_class_ids))
            if all_class_ids:
                qs = qs.filter(user__class_obj_id__in=all_class_ids)
        else:
            # LV2（评审老师）及更低：优先按分配任务 + scope 班级过滤
            assigned_submission_ids = list(
                ReviewAssignment.objects.filter(
                    reviewer=user, status='assigned'
                ).values_list('submission_id', flat=True).distinct()
            )
            from users.models import UserRole as URScope
            scope_class_ids = list(
                URScope.objects.filter(
                    user=user, scope_type='class', scope_id__isnull=False
                ).values_list('scope_id', flat=True)
            )
            if assigned_submission_ids or scope_class_ids:
                from django.db.models import Q as _Q
                sub_filter = _Q()
                if assigned_submission_ids:
                    sub_filter |= _Q(id__in=assigned_submission_ids)
                if scope_class_ids:
                    sub_filter |= _Q(user__class_obj_id__in=scope_class_ids)
                qs = qs.filter(sub_filter)
            else:
                # 无分配也无 scope 时，按最高角色 level 决定监督可见范围
                # （仅影响列表可见性，操作权限由各接口独立校验）
                max_level = _user_max_role_level(user)
                if max_level >= ROLE_LEVEL_SUPERADMIN:
                    pass
                elif max_level >= ROLE_LEVEL_DIRECTOR:
                    from users.models import UserRole as URDir
                    from org.models import Class as OrgClassDir
                    dept_ids = list(
                        URDir.objects.filter(user=user, scope_type='department')
                        .values_list('scope_id', flat=True)
                    )
                    cls_from_dept = list(
                        OrgClassDir.objects.filter(department_id__in=dept_ids)
                        .values_list('id', flat=True)
                    ) if dept_ids else []
                    cls_direct = list(
                        URDir.objects.filter(user=user, scope_type='class', scope_id__isnull=False)
                        .values_list('scope_id', flat=True)
                    )
                    all_cls = list(set(cls_from_dept + cls_direct))
                    if all_cls:
                        qs = qs.filter(user__class_obj_id__in=all_cls)
                else:
                    qs = qs.none()

        from django.db.models import Count, Q, Exists, OuterRef
        from users.models import UserRole as _UR
        qs = qs.annotate(
            _reviewer_count=Count(
                'score_records__reviewer',
                filter=Q(score_records__round_type__in=[1, 2]),
                distinct=True,
            ),
            _arb_count=Count('arbitration_records'),
            _is_assistant_sub=Exists(
                _UR.objects.filter(
                    user_id=OuterRef('user_id'),
                    role__level=ROLE_LEVEL_ASSISTANT,
                )
            ),
        )

        category = self.request.query_params.get('category', 'all')
        if category == 'disputed':
            qs = qs.filter(
                Q(_reviewer_count__gte=2) | Q(_arb_count__gt=0) | Q(_is_assistant_sub=True)
            )

        return qs


class ReviewSubmissionDetailAPIView(RetrieveAPIView):
    """GET /api/v1/review/submissions/<id>/ 某提交的审核视图。"""
    permission_classes = [IsAuthenticated]
    serializer_class = StudentSubmissionSerializer
    queryset = StudentSubmission.objects.all().select_related('project', 'user', 'user__class_obj', 'user__department').prefetch_related(
        Prefetch('evidences', queryset=Evidence.objects.filter(is_deleted=False).order_by('id')),
        'score_records',
    )

    def retrieve(self, request, *args, **kwargs):
        instance = self.get_object()
        if _project_has_assignments(instance.project_id):
            assignment = _get_active_assignment_for_submission(request.user, instance)
            if assignment is None and not user_level_at_least(request.user, ROLE_LEVEL_COUNSELOR):
                readable_assignment = _get_readable_assignment_for_submission(request.user, instance)
                if readable_assignment is None:
                    return Response({'detail': '您未被分配该提交的审核任务'}, status=status.HTTP_403_FORBIDDEN)
        response = super().retrieve(request, *args, **kwargs)
        from users.models import UserRole as _DetailUR
        response.data['is_assistant_submission'] = _DetailUR.objects.filter(
            user_id=instance.user_id,
            role__level=ROLE_LEVEL_ASSISTANT,
        ).exists()
        return response


class ReviewQuestionListAPIView(APIView):
    """GET /api/v1/review/submissions/<id>/questions/ 审核题目化视图。"""
    permission_classes = [IsAuthenticated]

    def get(self, request, pk):
        if not user_level_at_least(request.user, ROLE_LEVEL_ASSISTANT):
            return Response({'detail': '无权限查看审核题目'}, status=status.HTTP_403_FORBIDDEN)
        try:
            sub = StudentSubmission.objects.select_related('project').get(pk=pk)
        except StudentSubmission.DoesNotExist:
            return Response({'detail': '提交不存在'}, status=status.HTTP_404_NOT_FOUND)
        # 若项目启用分配任务，则必须命中任务归属
        if _project_has_assignments(sub.project_id):
            assignment = _get_active_assignment_for_submission(request.user, sub)
            if assignment is None and not user_level_at_least(request.user, ROLE_LEVEL_COUNSELOR):
                readable_assignment = _get_readable_assignment_for_submission(request.user, sub)
                if readable_assignment is None:
                    return Response({'detail': '您未被分配该提交的审核任务'}, status=status.HTTP_403_FORBIDDEN)

        indicators = list(
            EvalIndicator.objects.filter(project=sub.project, score_source='self').filter(
                Q(is_record_only=False) | Q(is_record_only=True, record_only_requires_review=True)
            )
            .select_related('parent', 'parent__parent')
            .order_by('order', 'id')
        )
        children_map = {}
        for ind in indicators:
            if ind.parent_id is not None:
                children_map.setdefault(ind.parent_id, []).append(ind.id)
        leaf_indicators = [ind for ind in indicators if ind.id not in children_map]
        indicator_ids = [i.id for i in leaf_indicators]

        answer_map = {
            ans.indicator_id: ans
            for ans in SubmissionAnswer.objects.filter(submission=sub, indicator_id__in=indicator_ids)
        }
        score_map = {}
        for rec in ScoreRecord.objects.filter(submission=sub, indicator_id__in=indicator_ids, reviewer=request.user):
            score_map[rec.indicator_id] = rec

        evidences = Evidence.objects.filter(
            submission=sub,
            indicator_id__in=indicator_ids,
            is_deleted=False,
        ).order_by('indicator_id', '-id')
        evidence_map = {}
        for ev in evidences:
            _, ext = os.path.splitext((ev.file.name if ev.file else '') or '')
            evidence_map.setdefault(ev.indicator_id, []).append({
                'id': ev.id,
                'name': ev.name or '',
                'file': ev.file.url if ev.file else '',
                'file_ext': ext.lower().lstrip('.'),
                'created_at': ev.created_at,
            })

        questions = []
        for ind in leaf_indicators:
            ans = answer_map.get(ind.id)
            my_score = score_map.get(ind.id)
            parent = ind.parent if ind.parent_id else None
            grandparent = parent.parent if (parent and parent.parent_id) else None
            if grandparent:
                root = grandparent
                section_name = grandparent.name
            elif parent:
                root = parent
                section_name = parent.name
            else:
                root = None
                section_name = ''
            ev_items = evidence_map.get(ind.id, [])
            questions.append({
                'indicator_id': ind.id,
                'indicator_name': ind.name,
                'section_name': section_name,
                'order': ind.order,
                'max_score': _raw_max_score(ind),
                'self_score': ans.self_score if ans else None,
                'process_record': (ans.process_record or '') if ans else '',
                'is_completed': ans.is_completed if ans else False,
                'evidence_count': len(ev_items),
                'require_process_record': ind.require_process_record,
                'is_record_only': bool(ind.is_record_only),
                'evidences': ev_items,
                'latest_score': my_score.score if my_score else None,
                'latest_comment': my_score.comment if my_score else '',
                'parent_indicator_id': ind.parent_id,
                'parent_agg_formula': parent.agg_formula if parent else None,
                'parent_max_score': str(parent.max_score) if parent and parent.max_score is not None else None,
                '_root_order': root.order if root else 0,
                '_root_id': root.id if root else 0,
                '_sub_order': parent.order if grandparent else 0,
                '_sub_id': parent.id if grandparent else 0,
            })
        questions.sort(key=lambda q: (q['_root_order'], q['_root_id'], q['_sub_order'], q['_sub_id'], q['order'], q['indicator_id']))
        return Response(ReviewQuestionSerializer(questions, many=True).data)


class ReviewObjectionCreateAPIView(APIView):
    """POST /api/v1/review/submissions/<id>/objections/ 发起模块异议上报。"""
    permission_classes = [IsAuthenticated]
    parser_classes = [MultiPartParser, FormParser]

    def post(self, request, pk):
        if not user_level_at_least(request.user, ROLE_LEVEL_ASSISTANT):
            return Response({'detail': '无权限发起异议上报'}, status=status.HTTP_403_FORBIDDEN)
        try:
            sub = StudentSubmission.objects.select_related('project', 'user').get(pk=pk)
        except StudentSubmission.DoesNotExist:
            return Response({'detail': '提交不存在'}, status=status.HTTP_404_NOT_FOUND)
        if sub.user_id == request.user.id:
            return Response({'detail': '不可对自己的提交发起异议'}, status=status.HTTP_403_FORBIDDEN)
        if sub.status not in ('submitted', 'under_review', 'approved'):
            return Response({'detail': '当前提交状态不允许发起异议'}, status=status.HTTP_400_BAD_REQUEST)

        indicator_id = request.data.get('indicator_id')
        reason = (request.data.get('reason') or '').strip()
        if not indicator_id or not reason:
            return Response({'detail': '请提供 indicator_id 和 reason'}, status=status.HTTP_400_BAD_REQUEST)
        try:
            ind = EvalIndicator.objects.get(pk=indicator_id, project=sub.project, score_source='self')
        except EvalIndicator.DoesNotExist:
            return Response({'detail': '仅可对评审模块发起异议上报'}, status=status.HTTP_400_BAD_REQUEST)

        level = request.user.current_role.level if request.user.current_role else -1
        source_role = 'assistant'
        assigned_to = None
        status_code = 'pending_counselor'
        handler_level = ROLE_LEVEL_COUNSELOR

        if level < ROLE_LEVEL_COUNSELOR:
            assignment = _get_active_assignment_for_submission(request.user, sub)
            if assignment is None or assignment.role_type != 'assistant':
                return Response({'detail': '您未被分配该提交的助理评审任务'}, status=status.HTTP_403_FORBIDDEN)
            class_id = getattr(sub.user, 'class_obj_id', None)
            if not class_id:
                return Response({'detail': '该学生未绑定班级，无法上报'}, status=status.HTTP_400_BAD_REQUEST)
            from .assignment_services import _find_counselor_owner
            assigned_to = _find_counselor_owner(class_id)
            if assigned_to is None:
                return Response({'detail': '未找到该班级对应评审老师，无法上报'}, status=status.HTTP_400_BAD_REQUEST)
        elif level == ROLE_LEVEL_COUNSELOR:
            if not _user_manages_submission_class(request.user, sub):
                assignment = _get_active_assignment_for_submission(request.user, sub)
                if assignment is None or assignment.role_type not in {'counselor', 'counselor_confirm', 'counselor_dispatch'}:
                    return Response({'detail': '您不在该提交管辖范围，无法上报异议'}, status=status.HTTP_403_FORBIDDEN)
            source_role = 'counselor'
            status_code = 'escalated_to_director'
            handler_level = ROLE_LEVEL_DIRECTOR
            assigned_to = _get_director_for_submission(sub)
            if assigned_to is None:
                return Response({'detail': '未找到该院系主任，无法上报'}, status=status.HTTP_400_BAD_REQUEST)
        else:
            return Response({'detail': '当前角色无需通过异议上报，请直接裁定'}, status=status.HTTP_400_BAD_REQUEST)

        objection = ReviewObjection.objects.create(
            submission=sub,
            indicator=ind,
            raised_by=request.user,
            assigned_to=assigned_to,
            source_role=source_role,
            status=status_code,
            current_handler_level=handler_level,
            reason=reason,
        )
        files = request.FILES.getlist('files')
        for f in files:
            ReviewObjectionAttachment.objects.create(
                objection=objection,
                file=f,
                name=getattr(f, 'name', '') or '',
                uploaded_by=request.user,
            )
        return Response(ReviewObjectionSerializer(objection).data, status=status.HTTP_201_CREATED)


class ReviewObjectionListAPIView(ListAPIView):
    """GET /api/v1/review/objections/ 审核异议工单列表。"""
    permission_classes = [IsAuthenticated]
    serializer_class = ReviewObjectionSerializer

    def get_queryset(self):
        user = self.request.user
        level = user.current_role.level if user.current_role else -1
        qs = ReviewObjection.objects.select_related(
            'submission', 'submission__user', 'submission__project', 'indicator', 'raised_by', 'assigned_to'
        ).prefetch_related('attachments').order_by('-created_at')

        if user_is_admin(user):
            pass
        elif level >= ROLE_LEVEL_DIRECTOR:
            dept_id = getattr(user, 'department_id', None)
            dept_filter = Q(pk__in=[])
            if dept_id:
                dept_filter = Q(submission__user__department_id=dept_id)
            qs = qs.filter(
                Q(assigned_to=user) |
                Q(raised_by=user) |
                dept_filter
            )
        elif level >= ROLE_LEVEL_COUNSELOR:
            class_ids = _get_counselor_class_ids(user)
            qs = qs.filter(
                Q(assigned_to=user) |
                Q(raised_by=user) |
                Q(submission__user__class_obj_id__in=class_ids)
            )
        else:
            qs = qs.filter(raised_by=user)

        status_filter = self.request.query_params.get('status')
        if status_filter:
            qs = qs.filter(status=status_filter)
        only_mine = self.request.query_params.get('mine')
        if only_mine in ('1', 'true'):
            qs = qs.filter(raised_by=user)
        return qs


class ReviewObjectionDetailAPIView(RetrieveAPIView):
    """GET /api/v1/review/objections/<id>/ 审核异议工单详情。"""
    permission_classes = [IsAuthenticated]
    serializer_class = ReviewObjectionSerializer
    queryset = ReviewObjection.objects.select_related(
        'submission', 'submission__user', 'submission__project', 'indicator', 'raised_by', 'assigned_to'
    ).prefetch_related('attachments')

    def retrieve(self, request, *args, **kwargs):
        obj = self.get_object()
        if not _can_user_access_objection(request.user, obj):
            return Response({'detail': '无权限查看该异议工单'}, status=status.HTTP_403_FORBIDDEN)
        return super().retrieve(request, *args, **kwargs)


class ReviewObjectionHandleAPIView(APIView):
    """POST /api/v1/review/objections/<id>/handle/ 处理异议工单。"""
    permission_classes = [IsAuthenticated]

    def post(self, request, pk):
        try:
            objection = ReviewObjection.objects.select_related('submission', 'submission__user', 'indicator').get(pk=pk)
        except ReviewObjection.DoesNotExist:
            return Response({'detail': '异议工单不存在'}, status=status.HTTP_404_NOT_FOUND)
        if not _can_user_access_objection(request.user, objection):
            return Response({'detail': '无权限处理该异议工单'}, status=status.HTTP_403_FORBIDDEN)

        level = request.user.current_role.level if request.user.current_role else -1
        action_type = request.data.get('action')  # resolve|reject|escalate_director|escalate_admin
        comment = (request.data.get('resolution_comment') or '').strip()
        raw_score = request.data.get('resolved_score')

        if objection.status == 'pending_counselor':
            if level < ROLE_LEVEL_COUNSELOR or (objection.assigned_to_id and objection.assigned_to_id != request.user.id and not user_is_admin(request.user)):
                return Response({'detail': '仅对应评审老师可处理该工单'}, status=status.HTTP_403_FORBIDDEN)
            if action_type == 'escalate_director':
                director = _get_director_for_submission(objection.submission)
                if director is None:
                    return Response({'detail': '未找到院系主任，无法上报'}, status=status.HTTP_400_BAD_REQUEST)
                objection.status = 'escalated_to_director'
                objection.current_handler_level = ROLE_LEVEL_DIRECTOR
                objection.assigned_to = director
                objection.resolution_comment = comment
                objection.save(update_fields=['status', 'current_handler_level', 'assigned_to', 'resolution_comment', 'updated_at'])
                return Response(ReviewObjectionSerializer(objection).data)
            if action_type not in ('resolve', 'reject'):
                return Response({'detail': 'action 须为 resolve、reject 或 escalate_director'}, status=status.HTTP_400_BAD_REQUEST)
        elif objection.status == 'escalated_to_director':
            if level < ROLE_LEVEL_DIRECTOR or (objection.assigned_to_id and objection.assigned_to_id != request.user.id and not user_is_admin(request.user)):
                return Response({'detail': '仅对应院系主任可处理该工单'}, status=status.HTTP_403_FORBIDDEN)
            if action_type == 'escalate_admin':
                admin_user = request.user if user_is_admin(request.user) else None
                if admin_user is None:
                    from users.models import UserRole
                    admin_ur = UserRole.objects.filter(role__level=ROLE_LEVEL_SUPERADMIN).select_related('user').first()
                    admin_user = admin_ur.user if admin_ur else None
                if admin_user is None:
                    return Response({'detail': '未找到超级管理员，无法上报'}, status=status.HTTP_400_BAD_REQUEST)
                objection.status = 'escalated_to_admin'
                objection.current_handler_level = ROLE_LEVEL_SUPERADMIN
                objection.assigned_to = admin_user
                objection.resolution_comment = comment
                objection.save(update_fields=['status', 'current_handler_level', 'assigned_to', 'resolution_comment', 'updated_at'])
                return Response(ReviewObjectionSerializer(objection).data)
            if action_type not in ('resolve', 'reject'):
                return Response({'detail': 'action 须为 resolve、reject 或 escalate_admin'}, status=status.HTTP_400_BAD_REQUEST)
        elif objection.status == 'escalated_to_admin':
            if not user_is_admin(request.user):
                return Response({'detail': '仅超级管理员可处理该工单'}, status=status.HTTP_403_FORBIDDEN)
            if action_type not in ('resolve', 'reject'):
                return Response({'detail': 'action 须为 resolve 或 reject'}, status=status.HTTP_400_BAD_REQUEST)
        else:
            return Response({'detail': '该异议工单已处理完成'}, status=status.HTTP_400_BAD_REQUEST)

        if action_type == 'reject':
            objection.status = 'rejected'
            objection.resolution_comment = comment
            objection.assigned_to = request.user
            objection.save(update_fields=['status', 'resolution_comment', 'assigned_to', 'updated_at'])
            return Response(ReviewObjectionSerializer(objection).data)

        score, score_error = _validate_indicator_score(objection.indicator, raw_score)
        if score_error:
            return Response({'detail': score_error}, status=status.HTTP_400_BAD_REQUEST)
        current_max_level = _user_max_role_level(request.user)
        existing_arb = ArbitrationRecord.objects.filter(
            submission=objection.submission,
            indicator=objection.indicator,
        ).first()
        if existing_arb and existing_arb.arbitrator_level is not None and existing_arb.arbitrator_level > current_max_level:
            return Response({'detail': '该指标已有更高权限仲裁记录，无法覆盖'}, status=status.HTTP_403_FORBIDDEN)

        ArbitrationRecord.objects.update_or_create(
            submission=objection.submission,
            indicator=objection.indicator,
            defaults={
                'arbitrator': request.user,
                'arbitrator_level': current_max_level,
                'score': score,
                'comment': comment,
                'triggered_reason': _triggered_reason_by_actor_level(current_max_level),
            },
        )
        recompute_submission_final_score(objection.submission)
        _maybe_finalize_submission_status(objection.submission)

        if level >= ROLE_LEVEL_SUPERADMIN:
            objection.status = 'resolved_by_admin'
        elif level >= ROLE_LEVEL_DIRECTOR:
            objection.status = 'resolved_by_director'
        else:
            objection.status = 'resolved_by_counselor'
        objection.resolution_comment = comment
        objection.resolved_score = score
        objection.assigned_to = request.user
        objection.save(update_fields=['status', 'resolution_comment', 'resolved_score', 'assigned_to', 'updated_at'])
        return Response(ReviewObjectionSerializer(objection).data)


class ReviewObjectionAttachmentUploadAPIView(APIView):
    """POST /api/v1/review/objections/<id>/attachments/ 上传异议附件。"""
    permission_classes = [IsAuthenticated]
    parser_classes = [MultiPartParser, FormParser]

    def post(self, request, pk):
        try:
            objection = ReviewObjection.objects.select_related('submission').get(pk=pk)
        except ReviewObjection.DoesNotExist:
            return Response({'detail': '异议工单不存在'}, status=status.HTTP_404_NOT_FOUND)
        if not _can_user_access_objection(request.user, objection):
            return Response({'detail': '无权限上传附件'}, status=status.HTTP_403_FORBIDDEN)
        files = request.FILES.getlist('files')
        if not files:
            return Response({'detail': '请至少上传一个附件'}, status=status.HTTP_400_BAD_REQUEST)
        created = []
        for f in files:
            att = ReviewObjectionAttachment.objects.create(
                objection=objection,
                file=f,
                name=getattr(f, 'name', '') or '',
                uploaded_by=request.user,
            )
            created.append(att)
        return Response(ReviewObjectionAttachmentSerializer(created, many=True).data, status=status.HTTP_201_CREATED)


class ReviewObjectionAttachmentDeleteAPIView(APIView):
    """DELETE /api/v1/review/objections/<id>/attachments/<aid>/ 删除异议附件。"""
    permission_classes = [IsAuthenticated]

    def delete(self, request, pk, attachment_id):
        try:
            objection = ReviewObjection.objects.get(pk=pk)
        except ReviewObjection.DoesNotExist:
            return Response({'detail': '异议工单不存在'}, status=status.HTTP_404_NOT_FOUND)
        if not _can_user_access_objection(request.user, objection):
            return Response({'detail': '无权限删除附件'}, status=status.HTTP_403_FORBIDDEN)
        try:
            att = ReviewObjectionAttachment.objects.get(pk=attachment_id, objection=objection)
        except ReviewObjectionAttachment.DoesNotExist:
            return Response({'detail': '附件不存在'}, status=status.HTTP_404_NOT_FOUND)
        att.delete()
        return Response({'detail': '删除成功'})


class ReviewInitialAPIView(APIView):
    """POST /api/v1/review/submissions/<id>/initial/ 评审老师初审（将状态改为 under_review）。"""
    permission_classes = [IsAuthenticated]

    def post(self, request, pk):
        # 评审老师（辅导员）（level>=2）可初审
        if not user_level_at_least(request.user, 2):
            return Response({'detail': f'{get_role_display_name(ROLE_LEVEL_COUNSELOR)}及以上角色方可初审'}, status=status.HTTP_403_FORBIDDEN)
        try:
            sub = StudentSubmission.objects.get(pk=pk)
        except StudentSubmission.DoesNotExist:
            return Response({'detail': '提交不存在'}, status=status.HTTP_404_NOT_FOUND)
        if sub.via_late_channel:
            return Response({'detail': '补交提交需在待推送中心推送后方可进入审核'}, status=status.HTTP_400_BAD_REQUEST)
        if sub.status != 'submitted':
            return Response({'detail': '仅已提交状态可初审'}, status=status.HTTP_400_BAD_REQUEST)
        sub.status = 'under_review'
        sub.save(update_fields=['status'])
        log_action(
            user=request.user,
            action='initial_review',
            module=OperationLog.MODULE_SCORING,
            level=OperationLog.LEVEL_NOTICE,
            target_type='submission',
            target_id=sub.id,
            target_repr=str(sub.id),
            request=request,
        )
        return Response(StudentSubmissionSerializer(sub).data)


class ReviewScoresListCreateAPIView(APIView):
    """GET /api/v1/review/submissions/<id>/scores/ 可见评分列表；POST 提交评分（双评）。"""
    permission_classes = [IsAuthenticated]

    def get(self, request, pk):
        try:
            sub = StudentSubmission.objects.get(pk=pk)
        except StudentSubmission.DoesNotExist:
            return Response({'detail': '提交不存在'}, status=status.HTTP_404_NOT_FOUND)
        if _project_has_assignments(sub.project_id):
            assignment = _get_active_assignment_for_submission(request.user, sub)
            if assignment is None and not user_level_at_least(request.user, ROLE_LEVEL_COUNSELOR):
                readable_assignment = _get_readable_assignment_for_submission(request.user, sub)
                if readable_assignment is None:
                    return Response({'detail': '您未被分配该提交的查看权限'}, status=status.HTTP_403_FORBIDDEN)
        rule = getattr(sub.project, 'review_rule', None)
        allow_other = rule.allow_view_other_scores if rule else False
        qs = ScoreRecord.objects.filter(submission=sub).select_related('indicator', 'reviewer')
        if not allow_other and not user_level_at_least(request.user, ROLE_LEVEL_COUNSELOR):
            qs = qs.filter(reviewer=request.user)
        score_data = ScoreRecordSerializer(qs, many=True).data

        arb_qs = ArbitrationRecord.objects.filter(submission=sub).select_related('indicator', 'arbitrator')
        for arb in arb_qs:
            score_data.append({
                'id': f'arb_{arb.id}',
                'submission': sub.id,
                'indicator': arb.indicator_id,
                'indicator_name': arb.indicator.name if arb.indicator else '',
                'reviewer': arb.arbitrator_id,
                'reviewer_name': arb.arbitrator.username if arb.arbitrator else None,
                'score': str(arb.score),
                'comment': arb.comment or '',
                'round_type': 3,
                'score_channel': 'arbitration',
                'scorer_role_level': arb.arbitrator_level,
                'scorer_max_role_level': arb.arbitrator_level,
                'is_delegated': False,
                'created_at': arb.created_at.isoformat() if arb.created_at else None,
            })

        score_data.sort(key=lambda x: x.get('created_at') or '', reverse=False)
        return Response(_attach_logical_round_fields(score_data))

    def post(self, request, pk):
        if not user_level_at_least(request.user, ROLE_LEVEL_ASSISTANT):
            return Response({'detail': '无权限评分'}, status=status.HTTP_403_FORBIDDEN)
        try:
            sub = StudentSubmission.objects.get(pk=pk)
        except StudentSubmission.DoesNotExist:
            return Response({'detail': '提交不存在'}, status=status.HTTP_404_NOT_FOUND)
        if sub.user_id == request.user.id:
            return Response({'detail': '不可评审自己的提交'}, status=status.HTTP_403_FORBIDDEN)
        if sub.status == 'submitted' and sub.via_late_channel:
            return Response({'detail': '补交提交需在待推送中心推送后方可评分'}, status=status.HTTP_400_BAD_REQUEST)
        if sub.status not in ('submitted', 'under_review'):
            return Response({'detail': '当前状态不可评分'}, status=status.HTTP_400_BAD_REQUEST)
        rule = getattr(sub.project, 'review_rule', None)
        assignment = _get_active_assignment_for_submission(request.user, sub)
        if assignment is None:
            # 无分配任务 → 禁止通过标准评分接口评分，引导走仲裁通道
            return Response(
                {'detail': '您未被分配该提交的评审任务。如需介入评分，请使用仲裁接口。'},
                status=status.HTTP_403_FORBIDDEN,
            )
        else:
            allowed_role_types = {'assistant'}
            if _is_single_review(rule):
                allowed_role_types = {_single_allowed_role_type(rule)}
            else:
                mode = (rule.counselor_participation_mode if rule else 'arbitration_only') or 'arbitration_only'
                if mode == 'always_confirm':
                    allowed_role_types.update({'counselor_confirm', 'counselor'})
            if assignment.role_type not in allowed_role_types:
                return Response(
                    {'detail': f'双评模式下，{get_role_display_name(ROLE_LEVEL_COUNSELOR)}请通过仲裁接口评分，常规评分通道仅限{get_role_display_name(ROLE_LEVEL_ASSISTANT)}。'},
                    status=status.HTTP_403_FORBIDDEN,
                )
        from django.utils import timezone as tz
        proj = sub.project
        if proj.review_end_time and tz.now() > proj.review_end_time and not sub.via_late_channel:
            return Response({'detail': '成绩评定时间已截止，无法提交评分'}, status=status.HTTP_400_BAD_REQUEST)
        indicator_id = request.data.get('indicator_id')
        raw_score = request.data.get('score')
        comment = request.data.get('comment', '')
        round_type = assignment.round_type
        if indicator_id is None or raw_score is None:
            return Response({'detail': '缺少 indicator_id 或 score'}, status=status.HTTP_400_BAD_REQUEST)
        try:
            ind = EvalIndicator.objects.get(project=sub.project_id, pk=indicator_id, score_source='self')
        except EvalIndicator.DoesNotExist:
            return Response({'detail': '指标不存在或不在题目化审核范围'}, status=status.HTTP_404_NOT_FOUND)
        if not _indicator_requires_review(ind):
            return Response({'detail': '该指标为仅记录模块，不需要评审打分'}, status=status.HTTP_400_BAD_REQUEST)
        score, score_error = _validate_indicator_score(ind, raw_score)
        if score_error:
            return Response({'detail': score_error}, status=status.HTTP_400_BAD_REQUEST)
        if _is_single_review(rule) and _single_mode_conflict_exists(
            submission=sub,
            indicator=ind,
            reviewer_id=request.user.id,
            round_type=round_type,
        ):
            return Response({'detail': '单评模式下同一指标仅允许1条常规评分记录'}, status=status.HTTP_400_BAD_REQUEST)
        if ind.parent_id:
            parent_ind = ind.parent
            if parent_ind and parent_ind.agg_formula == 'sum_capped' and parent_ind.max_score is not None:
                from django.db.models import Sum
                sibling_sum = ScoreRecord.objects.filter(
                    submission=sub,
                    indicator__parent=parent_ind,
                    reviewer=request.user,
                    round_type=round_type,
                ).exclude(indicator=ind).aggregate(
                    total=Sum('score')
                )['total'] or Decimal('0')
                cap = Decimal(str(parent_ind.max_score))
                remaining = cap - sibling_sum
                if score > remaining:
                    return Response(
                        {'detail': f'封顶求和限制：父项满分 {cap}，其他子项已用 {sibling_sum}，剩余可用 {remaining}'},
                        status=status.HTTP_400_BAD_REQUEST,
                    )
        ctx = _build_score_context(request.user)
        rec, created = ScoreRecord.objects.update_or_create(
            submission=sub, indicator=ind, reviewer=request.user, round_type=round_type,
            defaults={
                'score': score,
                'comment': comment,
                'score_channel': 'assignment',
                **ctx,
            },
        )
        recompute_submission_final_score(sub)
        _maybe_create_single_review_confirm(sub, request.user, assignment)
        _maybe_mark_assignment_completed(sub, request.user, assignment)
        _maybe_finalize_submission_status(sub)
        log_action(
            user=request.user,
            action='score_submit',
            module=OperationLog.MODULE_SCORING,
            level=OperationLog.LEVEL_NOTICE,
            target_type='score_record',
            target_id=rec.id,
            target_repr=f'提交#{sub.id} · 指标#{ind.id}',
            extra={
                'score': str(score),
                'round_type': round_type,
                'score_channel': 'assignment',
                'logical_round_label': _format_logical_round_label(round_type),
                **ctx,
            },
            request=request,
        )
        record_payload = ScoreRecordSerializer(rec).data
        record_payload = _attach_logical_round_fields([record_payload])[0]
        return Response(record_payload, status=status.HTTP_201_CREATED if created else status.HTTP_200_OK)


class ReviewArbitrateAPIView(APIView):
    """POST /api/v1/review/submissions/<id>/arbitrate/ 仲裁打分。"""
    permission_classes = [IsAuthenticated]

    def post(self, request, pk):
        # 评审老师（辅导员）（level>=2）可仲裁
        if not user_level_at_least(request.user, 2):
            return Response({'detail': f'{get_role_display_name(ROLE_LEVEL_COUNSELOR)}及以上角色方可仲裁'}, status=status.HTTP_403_FORBIDDEN)
        try:
            sub = StudentSubmission.objects.get(pk=pk)
        except StudentSubmission.DoesNotExist:
            return Response({'detail': '提交不存在'}, status=status.HTTP_404_NOT_FOUND)
        rule = getattr(sub.project, 'review_rule', None)
        mode = (rule.counselor_participation_mode if rule else 'arbitration_only') or 'arbitration_only'
        single_review = _is_single_review(rule)
        single_mode = (rule.single_review_mode if rule else 'assistant_single') or 'assistant_single'
        if user_is_admin(request.user):
            pass
        elif user_level_at_least(request.user, ROLE_LEVEL_DIRECTOR):
            sub_dept_id = getattr(sub.user, 'department_id', None)
            user_dept_id = getattr(request.user, 'department_id', None)
            if not sub_dept_id or sub_dept_id != user_dept_id:
                return Response({'detail': '该提交不在您的管辖院系内，禁止仲裁'}, status=status.HTTP_403_FORBIDDEN)
        else:
            if not _user_manages_submission_class(request.user, sub):
                max_level = _user_max_role_level(request.user)
                if max_level >= ROLE_LEVEL_SUPERADMIN:
                    pass
                elif max_level >= ROLE_LEVEL_DIRECTOR:
                    sub_dept_id = getattr(sub.user, 'department_id', None)
                    from users.models import UserRole as URArb
                    arb_dept_ids = list(
                        URArb.objects.filter(user=request.user, scope_type='department')
                        .values_list('scope_id', flat=True)
                    )
                    if not sub_dept_id or sub_dept_id not in arb_dept_ids:
                        return Response({'detail': '该提交不在您的管辖范围内，禁止仲裁'}, status=status.HTTP_403_FORBIDDEN)
                else:
                    return Response({'detail': '您不管辖该提交所属班级，禁止仲裁'}, status=status.HTTP_403_FORBIDDEN)
        # 成绩评定截止时间检查（补交通道提交不受此限制）
        from django.utils import timezone as tz
        proj = sub.project
        if proj.review_end_time and tz.now() > proj.review_end_time and not sub.via_late_channel:
            return Response({'detail': '成绩评定时间已截止，无法提交仲裁评分'}, status=status.HTTP_400_BAD_REQUEST)
        indicator_id = request.data.get('indicator_id')
        raw_score = request.data.get('score')
        comment = request.data.get('comment', '')
        if indicator_id is None or raw_score is None:
            return Response({'detail': '缺少 indicator_id 或 score'}, status=status.HTTP_400_BAD_REQUEST)
        try:
            ind = EvalIndicator.objects.get(project=sub.project_id, pk=indicator_id, score_source='self')
        except EvalIndicator.DoesNotExist:
            return Response({'detail': '指标不存在或不在题目化审核范围'}, status=status.HTTP_404_NOT_FOUND)
        if not _indicator_requires_review(ind):
            return Response({'detail': '该指标为仅记录模块，不需要仲裁评分'}, status=status.HTTP_400_BAD_REQUEST)
        score, score_error = _validate_indicator_score(ind, raw_score)
        if score_error:
            return Response({'detail': score_error}, status=status.HTTP_400_BAD_REQUEST)
        # 仅在阈值规则命中时允许仲裁（默认）。
        # 降权用户可主动仲裁，不受阈值限制（已有 is_delegated 审计标记）。
        ctx = _build_score_context(request.user)
        is_delegated_user = ctx.get('is_delegated', False)
        current_max_level = _user_max_role_level(request.user)
        is_counselor_in_dual_arb = (
            not single_review
            and mode == 'arbitration_only'
            and user_level_at_least(request.user, ROLE_LEVEL_COUNSELOR)
        )
        if rule and rule.require_arbitration_above_threshold and not (single_review and single_mode == 'assistant_single') and not is_delegated_user and not is_counselor_in_dual_arb:
            hit, reason = _threshold_hit_for_indicator(sub, ind, rule)
            if reason == '评分记录不足2条':
                return Response({'detail': '当前评分数不足，暂不可仲裁'}, status=status.HTTP_400_BAD_REQUEST)
            if reason == 'threshold_not_hit':
                return Response({'detail': '分差未超过阈值，暂不需要仲裁'}, status=status.HTTP_400_BAD_REQUEST)
        existing_arb = ArbitrationRecord.objects.filter(submission=sub, indicator=ind).first()
        if existing_arb and existing_arb.arbitrator_level is not None:
            if existing_arb.arbitrator_level > current_max_level:
                return Response(
                    {'detail': '该指标已有更高权限的仲裁记录，您无权覆盖'},
                    status=status.HTTP_403_FORBIDDEN,
                )
        arb, _ = ArbitrationRecord.objects.update_or_create(
            submission=sub, indicator=ind,
            defaults={
                'arbitrator': request.user,
                'arbitrator_level': current_max_level,
                'score': score,
                'comment': comment,
                'triggered_reason': (
                    'counselor_or_director_decision'
                    if single_review
                    else ('score_diff_exceeded' if mode == 'arbitration_only' else 'counselor_or_director_decision')
                ),
            },
        )
        recompute_submission_final_score(sub)
        sub.refresh_from_db(fields=['final_score'])
        if sub.final_score is not None and sub.status in ('submitted', 'under_review'):
            sub.status = 'approved'
            sub.save(update_fields=['status'])
            from django.utils import timezone as _tz
            ReviewAssignment.objects.filter(
                submission=sub,
                status='assigned',
            ).update(status='completed', completed_at=_tz.now())
        log_action(
            user=request.user,
            action='arbitrate',
            module=OperationLog.MODULE_SCORING,
            level=OperationLog.LEVEL_WARNING,
            target_type='arbitration_record',
            target_id=arb.id,
            target_repr=f'提交#{sub.id} · 指标#{ind.id}',
            extra={'score': str(score), **ctx},
            request=request,
        )
        from .serializers import ArbitrationRecordSerializer  # noqa: F811
        return Response(ArbitrationRecordSerializer(arb).data, status=status.HTTP_201_CREATED)


def _check_confirm_token(user, token):
    """校验密码确认 token（由 /auth/verify-password/ 签发，有效期 5 分钟）。"""
    from django.core.cache import cache
    if not token:
        return False
    cache_key = f'confirm_token:{user.id}:{token}'
    result = cache.get(cache_key)
    if result:
        cache.delete(cache_key)
    return bool(result)


class ReviewBatchArbitrateAPIView(APIView):
    """
    POST /api/v1/review/submissions/<id>/batch-arbitrate/
    整套仲裁：一次性为提交的所有指标批量写入仲裁分，需密码确认 token。
    """
    permission_classes = [IsAuthenticated]

    def post(self, request, pk):
        if not user_level_at_least(request.user, ROLE_LEVEL_COUNSELOR):
            return Response(
                {'detail': f'{get_role_display_name(ROLE_LEVEL_COUNSELOR)}及以上角色方可仲裁'},
                status=status.HTTP_403_FORBIDDEN,
            )
        confirm_token = request.data.get('confirm_token')
        if not _check_confirm_token(request.user, confirm_token):
            return Response({'detail': '密码确认 token 无效或已过期，请重新输入密码'}, status=status.HTTP_403_FORBIDDEN)
        scores_list = request.data.get('scores')
        if not scores_list or not isinstance(scores_list, list):
            return Response({'detail': '缺少 scores 数组'}, status=status.HTTP_400_BAD_REQUEST)

        try:
            sub = StudentSubmission.objects.select_related('project', 'user').get(pk=pk)
        except StudentSubmission.DoesNotExist:
            return Response({'detail': '提交不存在'}, status=status.HTTP_404_NOT_FOUND)

        # 权限检查（与单条仲裁一致）
        if user_is_admin(request.user):
            pass
        elif user_level_at_least(request.user, ROLE_LEVEL_DIRECTOR):
            sub_dept_id = getattr(sub.user, 'department_id', None)
            user_dept_id = getattr(request.user, 'department_id', None)
            if not sub_dept_id or sub_dept_id != user_dept_id:
                return Response({'detail': '该提交不在您的管辖院系内，禁止仲裁'}, status=status.HTTP_403_FORBIDDEN)
        else:
            if not _user_manages_submission_class(request.user, sub):
                max_level = _user_max_role_level(request.user)
                if max_level >= ROLE_LEVEL_SUPERADMIN:
                    pass
                elif max_level >= ROLE_LEVEL_DIRECTOR:
                    sub_dept_id = getattr(sub.user, 'department_id', None)
                    from users.models import UserRole as URBatch
                    arb_dept_ids = list(
                        URBatch.objects.filter(user=request.user, scope_type='department')
                        .values_list('scope_id', flat=True)
                    )
                    if not sub_dept_id or sub_dept_id not in arb_dept_ids:
                        return Response({'detail': '该提交不在您的管辖范围内，禁止仲裁'}, status=status.HTTP_403_FORBIDDEN)
                else:
                    return Response({'detail': '您不管辖该提交所属班级，禁止仲裁'}, status=status.HTTP_403_FORBIDDEN)

        from django.utils import timezone as tz
        proj = sub.project
        if proj.review_end_time and tz.now() > proj.review_end_time and not sub.via_late_channel:
            return Response({'detail': '成绩评定时间已截止，无法提交仲裁评分'}, status=status.HTTP_400_BAD_REQUEST)

        ctx = _build_score_context(request.user)
        current_max_level = _user_max_role_level(request.user)
        rule = getattr(sub.project, 'review_rule', None)
        single_review = _is_single_review(rule)

        indicator_ids = [item.get('indicator_id') for item in scores_list if item.get('indicator_id') is not None]
        indicators = {
            ind.id: ind
            for ind in EvalIndicator.objects.filter(project=sub.project_id, pk__in=indicator_ids, score_source='self')
        }

        results = []
        skipped = []
        with transaction.atomic():
            for item in scores_list:
                ind_id = item.get('indicator_id')
                raw_score = item.get('score')
                comment = item.get('comment', '')
                if ind_id is None or raw_score is None:
                    continue
                ind = indicators.get(ind_id)
                if ind is None:
                    skipped.append({'indicator_id': ind_id, 'reason': '指标不存在'})
                    continue
                if not _indicator_requires_review(ind):
                    skipped.append({'indicator_id': ind_id, 'reason': '仅记录模块不需要仲裁评分'})
                    continue
                score, score_error = _validate_indicator_score(ind, raw_score)
                if score_error:
                    skipped.append({'indicator_id': ind_id, 'reason': score_error})
                    continue
                # 层级保护检查
                existing_arb = ArbitrationRecord.objects.filter(submission=sub, indicator=ind).first()
                if existing_arb and existing_arb.arbitrator_level is not None:
                    if existing_arb.arbitrator_level > current_max_level:
                        skipped.append({'indicator_id': ind_id, 'reason': '已有更高权限仲裁记录'})
                        continue
                arb, _ = ArbitrationRecord.objects.update_or_create(
                    submission=sub, indicator=ind,
                    defaults={
                        'arbitrator': request.user,
                        'arbitrator_level': current_max_level,
                        'score': score,
                        'comment': comment,
                        'triggered_reason': (
                            'counselor_or_director_decision'
                            if single_review
                            else 'batch_arbitration'
                        ),
                    },
                )
                results.append({'indicator_id': ind_id, 'score': str(score)})

            recompute_submission_final_score(sub)

        sub.refresh_from_db(fields=['final_score'])
        if sub.final_score is not None and sub.status in ('submitted', 'under_review'):
            sub.status = 'approved'
            sub.save(update_fields=['status'])
            from django.utils import timezone as _tz
            ReviewAssignment.objects.filter(
                submission=sub,
                status='assigned',
            ).update(status='completed', completed_at=_tz.now())

        log_action(
            user=request.user,
            action='batch_arbitrate',
            module=OperationLog.MODULE_SCORING,
            level=OperationLog.LEVEL_WARNING,
            target_type='student_submission',
            target_id=sub.id,
            target_repr=f'提交#{sub.id} 整套仲裁 {len(results)} 题',
            extra={
                'submitted_count': len(results),
                'skipped_count': len(skipped),
                'skipped': skipped,
                'scores': results,
                **ctx,
            },
            request=request,
        )
        return Response({
            'detail': f'整套仲裁完成，已提交 {len(results)} 题' + (f'，跳过 {len(skipped)} 题' if skipped else ''),
            'submitted': results,
            'skipped': skipped,
        }, status=status.HTTP_201_CREATED)


class ScoringImportAPIView(APIView):
    """
    POST /api/v1/scoring/import/
    Excel 批量导入成绩（每生一行，列头动态识别指标）。

    支持两种调用方式：
    1. 携带 preview_token（推荐）：从缓存取预检快照，跳过重新解析 Excel，效率更高且数据一致。
    2. 直接上传 file（降级）：兼容老调用方式，重新解析 Excel。

    均支持 excluded_rows 参数（JSON 数组字符串）跳过指定行号。
    权限：level >= 2（评审老师及以上）。
    """
    permission_classes = [IsAuthenticated]
    parser_classes = [MultiPartParser, FormParser]

    _PREVIEW_CACHE_PREFIX = 'score_import_preview_'
    _PREVIEW_TTL = 900  # 15 分钟

    @staticmethod
    def _preview_cache_key(user_id, token):
        return f'score_import_preview_{user_id}_{token}'

    def post(self, request):
        if not user_level_at_least(request.user, ROLE_LEVEL_COUNSELOR):
            return Response({'detail': f'{get_role_display_name(ROLE_LEVEL_COUNSELOR)}及以上可导入成绩'}, status=status.HTTP_403_FORBIDDEN)

        project_id = request.data.get('project_id')
        if not project_id:
            return Response({'detail': '请提供 project_id'}, status=status.HTTP_400_BAD_REQUEST)
        try:
            project = EvalProject.objects.get(pk=project_id)
        except EvalProject.DoesNotExist:
            return Response({'detail': '项目不存在'}, status=status.HTTP_404_NOT_FOUND)
        actor_scope = _resolve_import_actor_scope(request.user)
        if not _user_can_manage_project_import(request.user, project, actor_scope=actor_scope):
            return Response({'detail': '无权限导入该项目（仅可操作本院系/班级范围）'}, status=status.HTTP_403_FORBIDDEN)
        blocked, policy_detail = _is_import_blocked_by_policy(request.user, project, actor_scope)
        if blocked:
            log_action(
                user=request.user,
                action='import_blocked_by_policy',
                module=OperationLog.MODULE_SCORING,
                level=OperationLog.LEVEL_WARNING,
                target_type='eval_project',
                target_id=project.id,
                target_repr=project.name,
                reason=policy_detail,
                is_audit_event=True,
                request=request,
            )
            return Response({'detail': policy_detail}, status=status.HTTP_403_FORBIDDEN)

        cfg = project.import_config or {}
        student_field = str(cfg.get('student_field', 'student_no')).strip() or 'student_no'
        comment = str(cfg.get('comment', '批量导入')).strip() or '批量导入'
        import_round_type = 1

        # 解析 excluded_rows（JSON 数组字符串 "[3,7,...]"）
        try:
            excluded_rows = set(json.loads(request.data.get('excluded_rows', '[]') or '[]'))
        except (ValueError, TypeError):
            excluded_rows = set()

        # 优先使用 preview_token 从缓存取快照
        preview_token = request.data.get('preview_token', '').strip()
        file_name = 'unknown'
        data_rows = None
        header_row = None

        if preview_token:
            from django.core.cache import cache
            snapshot = cache.get(self._preview_cache_key(request.user.id, preview_token))
            if not snapshot or snapshot.get('user_id') != request.user.id or str(snapshot.get('project_id')) != str(project_id):
                return Response({'detail': 'preview_token 已过期或无效，请重新预检后导入'}, status=status.HTTP_400_BAD_REQUEST)
            file_name = snapshot.get('file_name', 'unknown')
            header_row = snapshot.get('header_row', [])
            data_rows = snapshot.get('data_rows', [])
            cache.delete(self._preview_cache_key(request.user.id, preview_token))
        else:
            file_obj = request.FILES.get('file')
            if not file_obj:
                return Response({'detail': '请提供 preview_token 或 file'}, status=status.HTTP_400_BAD_REQUEST)
            file_name = file_obj.name
            try:
                header_row, data_rows = _parse_excel_rows(file_obj)
            except ValueError as e:
                return Response({'detail': str(e)}, status=status.HTTP_400_BAD_REQUEST)

        batch = ImportedScoreBatch.objects.create(
            project=project, uploaded_by=request.user, file_name=file_name, status='processing',
        )
        errors = []
        try:
            importable_map = {
                ind.name: ind
                for ind in project.indicators.filter(score_source='import').order_by('order', 'id')
            }
            STUDENT_COL = 0
            col_to_indicator = {
                col_idx: importable_map[col_name]
                for col_idx, col_name in enumerate(header_row)
                if col_idx > 1 and col_name in importable_map
            }
            if not col_to_indicator:
                raise ValueError(
                    '未在表头中找到任何可统一导入的指标列。'
                    '请使用"下载导入模板"获取最新模板后再填写上传。'
                )

            batch.row_count = len(data_rows)
            affected_sub_ids = set()

            from django.contrib.auth import get_user_model
            UserModel = get_user_model()
            operator_level = actor_scope['effective_level']
            is_director = operator_level == ROLE_LEVEL_DIRECTOR
            is_counselor = operator_level == ROLE_LEVEL_COUNSELOR
            counselor_class_ids = actor_scope['class_ids'] if is_counselor else set()
            director_department_id = actor_scope['department_id'] if is_director else None

            for i, row in enumerate(data_rows):
                row_num = i + 2
                if not row:
                    continue
                # 跳过用户手动排除的行
                if row_num in excluded_rows:
                    continue
                student_identifier = str(row[STUDENT_COL]).strip() if len(row) > STUDENT_COL and row[STUDENT_COL] is not None else ''
                if not student_identifier or student_identifier in ('（填入学号）', '示例学号'):
                    continue

                if student_field == 'username':
                    target_user = UserModel.objects.filter(username=student_identifier).first()
                else:
                    target_user = (
                        UserModel.objects.filter(student_no=student_identifier).first()
                        or UserModel.objects.filter(username=student_identifier).first()
                    )
                if not target_user:
                    errors.append({'row': row_num, 'message': f'用户不存在: {student_identifier}'})
                    continue
                # 院系主任行级权限
                if is_director:
                    if not director_department_id or target_user.department_id != director_department_id:
                        dept_name = getattr(target_user.department, 'name', '未知院系') if target_user.department_id else '未知院系'
                        errors.append({
                            'row': row_num,
                            'message': f'学号 {student_identifier} 所在院系（{dept_name}）不在您的管理范围，已跳过',
                        })
                        continue
                # 评审老师行级权限
                if is_counselor:
                    if not target_user.class_obj_id or target_user.class_obj_id not in counselor_class_ids:
                        errors.append({
                            'row': row_num,
                            'message': f'学号 {student_identifier} 所在班级不在您的负责范围，已跳过',
                        })
                        continue
                sub = StudentSubmission.objects.filter(project=project, user=target_user).first()
                if not sub:
                    errors.append({'row': row_num, 'message': f'该生无提交记录: {student_identifier}'})
                    continue

                row_has_score = False
                for col_idx, ind in col_to_indicator.items():
                    raw_val = row[col_idx] if len(row) > col_idx else None
                    if raw_val is None or str(raw_val).strip() == '':
                        continue
                    score_val, score_error = _validate_indicator_score(ind, raw_val)
                    if score_error:
                        errors.append({'row': row_num, 'message': score_error})
                        continue
                    ImportedScoreDetail.objects.create(
                        batch=batch, submission=sub, indicator=ind, user=target_user, score=score_val, source='import',
                    )
                    import_ctx = _build_score_context(request.user)
                    ScoreRecord.objects.update_or_create(
                        submission=sub, indicator=ind, reviewer=request.user, round_type=import_round_type,
                        defaults={
                            'score': score_val,
                            'comment': comment,
                            'score_channel': 'import',
                            **import_ctx,
                        },
                    )
                    row_has_score = True

                if row_has_score:
                    affected_sub_ids.add(sub.id)

            for sub_id in affected_sub_ids:
                sub_obj = StudentSubmission.objects.get(pk=sub_id)
                recompute_submission_final_score(sub_obj)
                _maybe_finalize_submission_status(sub_obj)
            batch.status = 'completed'
            batch.error_log = errors
            batch.save(update_fields=['status', 'row_count', 'error_log'])
        except Exception as e:
            batch.status = 'failed'
            batch.error_log = errors + [{'message': str(e)}]
            batch.save(update_fields=['status', 'error_log'])
            return Response({'detail': str(e), 'batch': ImportedScoreBatchSerializer(batch).data}, status=status.HTTP_400_BAD_REQUEST)
        log_action(
            user=request.user,
            action='scores_import',
            module=OperationLog.MODULE_SCORING,
            level=OperationLog.LEVEL_WARNING,
            target_type='imported_score_batch',
            target_id=batch.id,
            target_repr=f'{project.name} · {file_name}',
            extra={
                'row_count': batch.row_count,
                'error_count': len(errors),
                'excluded_rows_count': len(excluded_rows),
                'used_preview_token': bool(preview_token),
                'operator_level': actor_scope['effective_level'],
                'operator_current_level': actor_scope['current_level'],
                'operator_max_level': actor_scope['max_level'],
                'project_id': project.id,
            },
            is_audit_event=True,
            request=request,
        )
        return Response(ImportedScoreBatchSerializer(batch).data, status=status.HTTP_201_CREATED)


class ScoringImportPrecheckAPIView(APIView):
    """
    POST /api/v1/scoring/import/precheck/
    导入预检：解析 Excel 文件并做全量验证，不写库。
    预检通过（无阻断性错误）时生成 preview_token，前端持此 token 调用
    /scoring/import/ 完成两阶段导入，缓存有效期 15 分钟。

    阻断性错误：学号不存在、超出权限范围、分数格式无效。
    非阻断警告：该学生在本项目暂无提交记录（提醒确认）。
    权限：level >= 2（评审老师及以上）。
    """
    permission_classes = [IsAuthenticated]
    parser_classes = [MultiPartParser, FormParser]

    _PREVIEW_TTL = 900  # 15 分钟

    def post(self, request):
        if not user_level_at_least(request.user, ROLE_LEVEL_COUNSELOR):
            return Response({'detail': f'{get_role_display_name(ROLE_LEVEL_COUNSELOR)}及以上可使用预检功能'}, status=status.HTTP_403_FORBIDDEN)
        project_id = request.data.get('project_id')
        file_obj = request.FILES.get('file')
        if not project_id or not file_obj:
            return Response({'detail': '请提供 project_id 和 file'}, status=status.HTTP_400_BAD_REQUEST)
        try:
            project = EvalProject.objects.get(pk=project_id)
        except EvalProject.DoesNotExist:
            return Response({'detail': '项目不存在'}, status=status.HTTP_404_NOT_FOUND)
        actor_scope = _resolve_import_actor_scope(request.user)
        if not _user_can_manage_project_import(request.user, project, actor_scope=actor_scope):
            return Response({'detail': '无权限预检该项目（仅可操作本院系/班级范围）'}, status=status.HTTP_403_FORBIDDEN)
        blocked, policy_detail = _is_import_blocked_by_policy(request.user, project, actor_scope)
        if blocked:
            log_action(
                user=request.user,
                action='import_blocked_by_policy',
                module=OperationLog.MODULE_SCORING,
                level=OperationLog.LEVEL_WARNING,
                target_type='eval_project',
                target_id=project.id,
                target_repr=project.name,
                reason=policy_detail,
                is_audit_event=True,
                request=request,
            )
            return Response({'detail': policy_detail}, status=status.HTTP_403_FORBIDDEN)

        cfg = project.import_config or {}
        student_field = str(cfg.get('student_field', 'student_no')).strip() or 'student_no'

        errors = []
        warnings = []
        total_rows = 0

        try:
            header_row, data_rows = _parse_excel_rows(file_obj)

            importable_map = {
                ind.name: ind
                for ind in project.indicators.filter(score_source='import').order_by('order', 'id')
            }
            STUDENT_COL = 0
            col_to_indicator = {
                col_idx: importable_map[col_name]
                for col_idx, col_name in enumerate(header_row)
                if col_idx > 1 and col_name in importable_map
            }
            if not col_to_indicator:
                return Response({
                    'total_rows': 0,
                    'errors': [{'row': None, 'message': '未在表头中找到任何可统一导入的指标列，请使用"下载导入模板"获取最新模板后再上传'}],
                    'warnings': [],
                    'has_errors': True,
                    'preview_token': None,
                })

            from django.contrib.auth import get_user_model
            UserModel = get_user_model()
            operator_level = actor_scope['effective_level']
            is_director = operator_level == ROLE_LEVEL_DIRECTOR
            is_counselor = operator_level == ROLE_LEVEL_COUNSELOR
            counselor_class_ids = actor_scope['class_ids'] if is_counselor else set()
            director_department_id = actor_scope['department_id'] if is_director else None

            for i, row in enumerate(data_rows):
                if not row:
                    continue
                student_identifier = str(row[STUDENT_COL]).strip() if len(row) > STUDENT_COL and row[STUDENT_COL] is not None else ''
                if not student_identifier or student_identifier in ('（填入学号）', '示例学号'):
                    continue
                total_rows += 1
                row_num = i + 2

                if student_field == 'username':
                    target_user = UserModel.objects.filter(username=student_identifier).first()
                else:
                    target_user = (
                        UserModel.objects.filter(student_no=student_identifier).first()
                        or UserModel.objects.filter(username=student_identifier).first()
                    )

                if not target_user:
                    errors.append({'row': row_num, 'message': f'学号/用户名不存在: {student_identifier}'})
                    continue

                # 院系主任行级权限检查
                if is_director:
                    if not director_department_id or target_user.department_id != director_department_id:
                        dept_name = getattr(target_user.department, 'name', '未知院系') if target_user.department_id else '未知院系'
                        errors.append({
                            'row': row_num,
                            'message': f'学号 {student_identifier}（{target_user.get_full_name() or target_user.username}）所在院系"{dept_name}"不在您的管理范围',
                        })
                        continue

                # 评审老师行级权限检查
                if is_counselor:
                    if not target_user.class_obj_id or target_user.class_obj_id not in counselor_class_ids:
                        errors.append({
                            'row': row_num,
                            'message': f'学号 {student_identifier}（{target_user.get_full_name() or target_user.username}）所在班级不在您的负责范围',
                        })
                        continue

                # 有无提交记录（非阻断）
                sub = StudentSubmission.objects.filter(project=project, user=target_user).first()
                if not sub:
                    warnings.append({
                        'row': row_num,
                        'message': f'学号 {student_identifier}（{target_user.get_full_name() or target_user.username}）在本项目暂无提交记录，导入时将跳过',
                    })
                    continue

                # 分数格式检查
                for col_idx, ind in col_to_indicator.items():
                    raw_val = row[col_idx] if len(row) > col_idx else None
                    if raw_val is None or str(raw_val).strip() == '':
                        continue
                    _, score_error = _validate_indicator_score(ind, raw_val)
                    if score_error:
                        errors.append({'row': row_num, 'message': score_error})

        except ValueError as e:
            return Response({
                'total_rows': 0,
                'errors': [{'row': None, 'message': str(e)}],
                'warnings': [],
                'has_errors': True,
                'preview_token': None,
            })
        except Exception as e:
            return Response({'detail': f'预检时发生错误: {e}'}, status=status.HTTP_400_BAD_REQUEST)

        # 无阻断错误时生成 preview_token，缓存行快照供后续导入直接使用
        preview_token = None
        if not errors:
            import uuid
            from django.core.cache import cache
            preview_token = uuid.uuid4().hex
            cache_key = f'score_import_preview_{request.user.id}_{preview_token}'
            cache.set(cache_key, {
                'user_id': request.user.id,
                'project_id': str(project_id),
                'file_name': getattr(file_obj, 'name', 'import.xlsx'),
                'header_row': header_row,
                'data_rows': [list(r) for r in data_rows],
            }, timeout=self._PREVIEW_TTL)

        return Response({
            'total_rows': total_rows,
            'errors': errors,
            'warnings': warnings,
            'has_errors': len(errors) > 0,
            'preview_token': preview_token,
        })


class ScoringImportTemplateAPIView(APIView):
    """
    GET /api/v1/scoring/import/template/?project_id=
    根据项目指标结构生成 Excel 导入模板。
    权限：level >= 2（评审老师及以上）。
    """
    permission_classes = [IsAuthenticated]

    def get(self, request):
        if not user_level_at_least(request.user, ROLE_LEVEL_COUNSELOR):
            return Response({'detail': f'{get_role_display_name(ROLE_LEVEL_COUNSELOR)}及以上可下载导入模板'}, status=status.HTTP_403_FORBIDDEN)
        project_id = request.query_params.get('project_id')
        if not project_id:
            return Response({'detail': '请提供 project_id'}, status=status.HTTP_400_BAD_REQUEST)
        try:
            project = EvalProject.objects.prefetch_related('indicators').get(pk=project_id)
        except EvalProject.DoesNotExist:
            return Response({'detail': '项目不存在'}, status=status.HTTP_404_NOT_FOUND)
        actor_scope = _resolve_import_actor_scope(request.user)
        if not _user_can_manage_project_import(request.user, project, actor_scope=actor_scope):
            return Response({'detail': '无权限下载该项目模板（仅可操作本院系/班级范围）'}, status=status.HTTP_403_FORBIDDEN)
        blocked, policy_detail = _is_import_blocked_by_policy(request.user, project, actor_scope)
        if blocked:
            return Response({'detail': policy_detail}, status=status.HTTP_403_FORBIDDEN)

        import openpyxl
        from openpyxl.styles import Font, PatternFill, Alignment
        from django.http import HttpResponse
        from io import BytesIO

        importable_indicators = [
            ind for ind in project.indicators.order_by('order', 'id')
            if ind.score_source == 'import'
        ]

        wb = openpyxl.Workbook()
        ws = wb.active
        ws.title = '成绩导入模板'

        header_row = ['学号', '姓名（可选）'] + [ind.name for ind in importable_indicators]
        ws.append(header_row)

        header_fill = PatternFill(start_color='D9E1F2', end_color='D9E1F2', fill_type='solid')
        for col_idx, _ in enumerate(header_row, start=1):
            cell = ws.cell(row=1, column=col_idx)
            cell.font = Font(bold=True)
            cell.fill = header_fill
            cell.alignment = Alignment(horizontal='center')

        example_row = ['（填入学号）', '（可空）'] + [0] * len(importable_indicators)
        ws.append(example_row)

        ws.column_dimensions['A'].width = 16
        ws.column_dimensions['B'].width = 12
        for col_idx in range(3, len(header_row) + 1):
            col_letter = openpyxl.utils.get_column_letter(col_idx)
            ws.column_dimensions[col_letter].width = max(14, len(header_row[col_idx - 1]) * 2 + 2)

        # 填写说明 sheet，注明操作者权限范围
        ws_note = wb.create_sheet('填写说明')
        ws_note.append(['字段', '说明'])
        ws_note.append(['学号', '学生学号（必填），系统按学号匹配学生'])
        ws_note.append(['姓名（可选）', '仅用于人工核对，不参与导入逻辑'])
        for ind in importable_indicators:
            parent_name = ind.parent.name if ind.parent_id else ''
            raw_ms = _raw_max_score(ind)
            ws_note.append([ind.name, f'所属：{parent_name}，满分：{"无限制" if raw_ms is None else raw_ms}'])
        ws_note.append(['', ''])
        ws_note.append(['注意', '请勿修改表头行（第1行）；从第2行起每行填写一名学生的成绩'])
        ws_note.append(['注意', '分数须为数字，不可填写文字'])
        ws_note.append(['', ''])

        # 根据操作者角色写入权限范围说明
        operator_level = actor_scope['effective_level']
        if operator_level == ROLE_LEVEL_DIRECTOR:
            dept_name = request.user.department.name if request.user.department_id else '（未绑定院系）'
            ws_note.append(['操作者权限范围', f'{get_role_display_name(ROLE_LEVEL_DIRECTOR)} — 仅可导入院系"{dept_name}"内学生的成绩'])
        elif operator_level == ROLE_LEVEL_COUNSELOR:
            from users.models import UserRole as UserRoleModel
            from org.models import Class as OrgClass
            scope_ids = list(
                UserRoleModel.objects.filter(
                    user=request.user, scope_type='class', scope_id__isnull=False
                ).values_list('scope_id', flat=True)
            )
            class_names = list(OrgClass.objects.filter(id__in=scope_ids).values_list('name', flat=True))
            classes_str = '、'.join(class_names) if class_names else '（未绑定班级）'
            ws_note.append(['操作者权限范围', f'{get_role_display_name(ROLE_LEVEL_COUNSELOR)} — 仅可导入负责班级（{classes_str}）内学生的成绩'])
        else:
            ws_note.append(['操作者权限范围', f'{get_role_display_name(ROLE_LEVEL_SUPERADMIN)} — 可导入所有学生的成绩'])

        buf = BytesIO()
        wb.save(buf)
        buf.seek(0)
        resp = HttpResponse(
            buf.getvalue(),
            content_type='application/vnd.openxmlformats-officedocument.spreadsheetml.sheet',
        )
        safe_name = project.name.replace(' ', '_')[:30]
        resp['Content-Disposition'] = f'attachment; filename="import_template_{safe_name}.xlsx"'
        return resp


# ---------------------------------------------------------------------------
# 学生助理（评卷助理）功能
# ---------------------------------------------------------------------------

class ReviewAssignmentGenerateAPIView(APIView):
    """
    POST /api/v1/review/assignments/generate/
    手动生成某项目的双评任务分配（按当前 review_rule 策略）。
    """
    permission_classes = [IsAuthenticated]

    def post(self, request):
        if not user_level_at_least(request.user, 2):
            return Response({'detail': '无权限生成分配任务'}, status=status.HTTP_403_FORBIDDEN)
        project_id = request.data.get('project_id')
        if not project_id:
            return Response({'detail': '请提供 project_id'}, status=status.HTTP_400_BAD_REQUEST)
        try:
            project = EvalProject.objects.get(pk=project_id)
        except EvalProject.DoesNotExist:
            return Response({'detail': '项目不存在'}, status=status.HTTP_404_NOT_FOUND)
        result = generate_review_assignments_for_project(project=project, operator_user=request.user)
        log_action(
            user=request.user,
            action='review_assignment_generate',
            module=OperationLog.MODULE_SCORING,
            level=OperationLog.LEVEL_WARNING,
            target_type='eval_project',
            target_id=project.id,
            target_repr=project.name,
            extra=result,
            is_audit_event=True,
            request=request,
        )
        return Response(result)


class MyReviewAssignmentsAPIView(ListAPIView):
    """GET /api/v1/review/assignments/my/ 查看我的评审分配任务。"""
    permission_classes = [IsAuthenticated]
    serializer_class = ReviewAssignmentSerializer

    def get_queryset(self):
        user = self.request.user
        project_id = self.request.query_params.get('project_id')
        qs = ReviewAssignment.objects.filter(reviewer=user, status='assigned').select_related(
            'project', 'submission', 'submission__user', 'reviewer'
        )
        if project_id:
            qs = qs.filter(project_id=project_id)
        return qs.order_by('-created_at')


class ProjectReviewAssignmentSummaryAPIView(APIView):
    """GET /api/v1/review/assignments/summary/?project_id= 查看项目分配概览。"""
    permission_classes = [IsAuthenticated]

    def get(self, request):
        if not user_level_at_least(request.user, 2):
            return Response({'detail': '无权限'}, status=status.HTTP_403_FORBIDDEN)
        project_id = request.query_params.get('project_id')
        if not project_id:
            return Response({'detail': '请提供 project_id'}, status=status.HTTP_400_BAD_REQUEST)
        qs = ReviewAssignment.objects.filter(project_id=project_id, status='assigned')
        by_role_type = {}
        by_round_type = {}
        for a in qs:
            by_role_type[a.role_type] = by_role_type.get(a.role_type, 0) + 1
            rt_key = str(a.round_type)
            by_round_type[rt_key] = by_round_type.get(rt_key, 0) + 1
        return Response({
            'project_id': int(project_id),
            'total': qs.count(),
            'by_role_type': by_role_type,
            'by_round_type': by_round_type,
        })


class CounselorAssignmentReleaseAPIView(APIView):
    """POST /api/v1/review/assignments/release/ 辅导员放行任务到助理。"""
    permission_classes = [IsAuthenticated]

    def post(self, request):
        if not user_level_at_least(request.user, 2):
            return Response({'detail': f'仅{get_role_display_name(ROLE_LEVEL_COUNSELOR)}及以上可放行任务'}, status=status.HTTP_403_FORBIDDEN)

        project_id = request.data.get('project_id')
        if not project_id:
            return Response({'detail': '请提供 project_id'}, status=status.HTTP_400_BAD_REQUEST)

        submission_ids = request.data.get('submission_ids') or []
        if submission_ids and not isinstance(submission_ids, list):
            return Response({'detail': 'submission_ids 必须为数组'}, status=status.HTTP_400_BAD_REQUEST)
        assistant_user_ids = request.data.get('assistant_user_ids') or []
        if assistant_user_ids and not isinstance(assistant_user_ids, list):
            return Response({'detail': 'assistant_user_ids 必须为数组'}, status=status.HTTP_400_BAD_REQUEST)
        reuse_latest = bool(request.data.get('reuse_latest'))
        force_reassign = bool(request.data.get('force_reassign'))

        try:
            project = EvalProject.objects.get(pk=project_id)
        except EvalProject.DoesNotExist:
            return Response({'detail': '项目不存在'}, status=status.HTTP_404_NOT_FOUND)
        rule = getattr(project, 'review_rule', None)
        if rule is None:
            return Response({'detail': '项目未配置评审规则'}, status=status.HTTP_400_BAD_REQUEST)

        dispatch_qs = ReviewAssignment.objects.filter(
            project_id=project_id,
            reviewer=request.user,
            role_type='counselor_dispatch',
            status='assigned',
        ).select_related('submission', 'submission__user', 'submission__user__class_obj')
        if submission_ids:
            dispatch_qs = dispatch_qs.filter(submission_id__in=submission_ids)
        dispatch_rows = list(dispatch_qs)
        if not dispatch_rows:
            return Response({'detail': '无可放行任务'}, status=status.HTTP_400_BAD_REQUEST)

        from django.contrib.auth import get_user_model
        from django.utils import timezone
        UserModel = get_user_model()
        fixed_assistants = list(UserModel.objects.filter(id__in=assistant_user_ids, is_active=True).order_by('id')) if assistant_user_ids else []

        created = 0
        released = 0
        skipped = []
        with transaction.atomic():
            for dispatch in dispatch_rows:
                sub = dispatch.submission
                if force_reassign:
                    ReviewAssignment.objects.filter(
                        project_id=project_id,
                        submission=sub,
                        role_type='assistant',
                        status='assigned',
                    ).update(status='cancelled')

                required_count = 1
                if rule.dual_review_enabled:
                    required_count = max(2, int(rule.allowed_assistant_count_per_submission or 2))

                if fixed_assistants:
                    candidates = _resolve_candidates_for_submission(sub, rule, request.user)
                    candidate_ids = {u.id for u in candidates}
                    selected = [u for u in fixed_assistants if u.id in candidate_ids and u.id != sub.user_id][:required_count]
                elif reuse_latest:
                    prev_reviewer_ids = (
                        ReviewAssignment.objects.filter(
                            project_id=project_id,
                            submission=sub,
                            role_type='assistant',
                            status__in=['completed', 'cancelled'],
                        )
                        .order_by('-created_at')
                        .values_list('reviewer_id', flat=True)
                        .distinct()
                    )
                    selected = list(
                        UserModel.objects.filter(
                            id__in=prev_reviewer_ids, is_active=True
                        ).order_by('id')
                    )[:required_count] if prev_reviewer_ids else []
                else:
                    candidates = _resolve_candidates_for_submission(sub, rule, request.user)
                    selected = _pick_assistants_for_submission(
                        submission=sub,
                        candidates=list(candidates),
                        count=required_count,
                        shuffle_enabled=bool(rule.cross_class_shuffle_enabled),
                    )

                if len(selected) < required_count:
                    skipped.append({'submission_id': sub.id, 'reason': '可用助理不足，未放行'})
                    continue

                round_no = 1
                for assistant in selected:
                    ReviewAssignment.objects.update_or_create(
                        submission=sub,
                        project_id=project_id,
                        reviewer=assistant,
                        role_type='assistant',
                        round_type=round_no,
                        defaults={
                            'strategy_mode': rule.review_scope_mode if rule.dual_review_enabled else 'assistant_single',
                            'assignment_version': 1,
                            'assigned_by': request.user,
                            'status': 'assigned',
                        },
                    )
                    round_no += 1
                    created += 1

                if rule.dual_review_enabled and rule.counselor_participation_mode == 'always_confirm':
                    ReviewAssignment.objects.update_or_create(
                        submission=sub,
                        project_id=project_id,
                        reviewer=request.user,
                        role_type='counselor_confirm',
                        round_type=4,
                        defaults={
                            'strategy_mode': rule.review_scope_mode,
                            'assignment_version': 1,
                            'assigned_by': request.user,
                            'status': 'assigned',
                        },
                    )

                dispatch.status = 'released'
                dispatch.completed_at = timezone.now()
                dispatch.save(update_fields=['status', 'completed_at', 'updated_at'])
                released += 1

        from realtime.registry import broadcast
        broadcast({'type': 'data_changed', 'model': 'review_assignment'})

        log_action(
            user=request.user,
            action='review_assignment_release',
            module=OperationLog.MODULE_SCORING,
            level=OperationLog.LEVEL_WARNING,
            target_type='eval_project',
            target_id=int(project_id),
            target_repr=project.name,
            extra={
                'released_count': released,
                'created_assistant_tasks': created,
                'skipped': skipped,
                'reuse_latest': reuse_latest,
                'force_reassign': force_reassign,
            },
            is_audit_event=True,
            request=request,
        )
        return Response({
            'detail': '放行完成',
            'released_count': released,
            'created_assistant_tasks': created,
            'skipped': skipped,
        })


class AssistantListAPIView(APIView):
    """
    GET /api/v1/review/assistant/list/?class_id=<id>
    查询某班级当前的学生助理列表（评审老师及以上可访问）。
    """
    permission_classes = [IsAuthenticated]

    def get(self, request):
        if not user_level_at_least(request.user, 2):
            return Response({'detail': '无权限'}, status=status.HTTP_403_FORBIDDEN)

        from django.contrib.auth import get_user_model
        from users.models import UserRole

        class_id = request.query_params.get('class_id')
        if not class_id:
            return Response({'detail': '请提供 class_id'}, status=status.HTTP_400_BAD_REQUEST)

        # 验证操作人是否管辖该班级（超级管理员除外）
        if not user_is_super_admin(request.user):
            counselor_manages = UserRole.objects.filter(
                user=request.user,
                scope_type='class',
                scope_id=int(class_id),
            ).exists()
            if not counselor_manages:
                return Response({'detail': '您不管辖该班级'}, status=status.HTTP_403_FORBIDDEN)

        User = get_user_model()
        assistant_user_ids = UserRole.objects.filter(
            role__level=ROLE_LEVEL_ASSISTANT,
            scope_type='class',
            scope_id=int(class_id),
        ).values_list('user_id', flat=True)

        from users.serializers import UserListSerializer
        assistants = User.objects.filter(id__in=assistant_user_ids).order_by('student_no', 'username')
        return Response({
            'class_id': class_id,
            'assistants': UserListSerializer(assistants, many=True).data,
        })


class AssistantAssignAPIView(APIView):
    """
    POST /api/v1/review/assistant/assign/
    评审老师（辅导员，level>=2）指派本班学生担任学生助理。

    请求体::

        {
            "student_id": 1,
            "class_id": 1
        }

    - student_id: 被指派的学生 ID
    - class_id: 关联班级 ID（必须是辅导员管辖的班级）
    """
    permission_classes = [IsAuthenticated]

    def post(self, request):
        if not user_level_at_least(request.user, 2):
            return Response({'detail': f'仅{get_role_display_name(ROLE_LEVEL_COUNSELOR)}及以上角色可指派{get_role_display_name(ROLE_LEVEL_ASSISTANT)}'}, status=status.HTTP_403_FORBIDDEN)

        from django.contrib.auth import get_user_model
        from users.models import UserRole, Role
        from org.models import Class

        student_id = request.data.get('student_id')
        class_id = request.data.get('class_id')
        if not student_id or not class_id:
            return Response({'detail': '请提供 student_id 和 class_id'}, status=status.HTTP_400_BAD_REQUEST)

        # 验证班级存在
        try:
            cls = Class.objects.get(pk=class_id)
        except Class.DoesNotExist:
            return Response({'detail': '班级不存在'}, status=status.HTTP_404_NOT_FOUND)

        # 验证操作人是否管辖该班级（超级管理员除外）
        if not user_is_super_admin(request.user):
            counselor_manages = UserRole.objects.filter(
                user=request.user,
                scope_type='class',
                scope_id=int(class_id),
            ).exists()
            if not counselor_manages:
                return Response({'detail': '您不管辖该班级，无法指派该班学生'}, status=status.HTTP_403_FORBIDDEN)

        # 验证被指派学生存在且属于该班级
        User = get_user_model()
        try:
            student = User.objects.get(pk=student_id)
        except User.DoesNotExist:
            return Response({'detail': '学生不存在'}, status=status.HTTP_404_NOT_FOUND)

        if student.class_obj_id != int(class_id):
            return Response({'detail': '该学生不属于指定班级，无法指派'}, status=status.HTTP_400_BAD_REQUEST)

        # 获取或创建 student_assistant 角色
        assistant_role = Role.objects.filter(level=ROLE_LEVEL_ASSISTANT).order_by('id').first()
        if not assistant_role:
            return Response({'detail': f'系统中未找到{get_role_display_name(ROLE_LEVEL_ASSISTANT)}角色，请联系{get_role_display_name(ROLE_LEVEL_SUPERADMIN)}'}, status=status.HTTP_500_INTERNAL_SERVER_ERROR)

        # 创建 UserRole 记录（scope 绑定到班级）
        ur, created = UserRole.objects.get_or_create(
            user=student,
            role=assistant_role,
            scope_id=int(class_id),
            scope_type='class',
            defaults={'is_primary': False},
        )

        log_action(
            user=request.user,
            action='assistant_assign',
            module=OperationLog.MODULE_USERS,
            level=OperationLog.LEVEL_WARNING,
            target_type='user',
            target_id=student.id,
            target_repr=student.username,
            extra={
                'class_id': class_id,
                'class_name': cls.name,
                'action': '指派' if created else '已存在（无变化）',
            },
            request=request,
        )

        return Response({
            'detail': '指派成功' if created else f'该学生已是{get_role_display_name(ROLE_LEVEL_ASSISTANT)}（无需重复指派）',
            'student_id': student.id,
            'class_id': class_id,
            'already_existed': not created,
        }, status=status.HTTP_201_CREATED if created else status.HTTP_200_OK)


class AssistantRevokeAPIView(APIView):
    """
    DELETE /api/v1/review/assistant/revoke/
    评审老师（辅导员，level>=2）撤销学生助理身份。

    请求体::

        {
            "student_id": 1,
            "class_id": 1
        }
    """
    permission_classes = [IsAuthenticated]

    def delete(self, request):
        if not user_level_at_least(request.user, 2):
            return Response({'detail': f'仅{get_role_display_name(ROLE_LEVEL_COUNSELOR)}及以上角色可撤销{get_role_display_name(ROLE_LEVEL_ASSISTANT)}'}, status=status.HTTP_403_FORBIDDEN)

        from django.contrib.auth import get_user_model
        from users.models import UserRole, Role
        from org.models import Class

        student_id = request.data.get('student_id')
        class_id = request.data.get('class_id')
        if not student_id or not class_id:
            return Response({'detail': '请提供 student_id 和 class_id'}, status=status.HTTP_400_BAD_REQUEST)

        # 验证操作人是否管辖该班级（超级管理员除外）
        if not user_is_super_admin(request.user):
            counselor_manages = UserRole.objects.filter(
                user=request.user,
                scope_type='class',
                scope_id=int(class_id),
            ).exists()
            if not counselor_manages:
                return Response({'detail': '您不管辖该班级，无法撤销'}, status=status.HTTP_403_FORBIDDEN)

        User = get_user_model()
        try:
            student = User.objects.get(pk=student_id)
        except User.DoesNotExist:
            return Response({'detail': '学生不存在'}, status=status.HTTP_404_NOT_FOUND)

        assistant_role = Role.objects.filter(level=ROLE_LEVEL_ASSISTANT).order_by('id').first()
        if not assistant_role:
            return Response({'detail': f'系统中未找到{get_role_display_name(ROLE_LEVEL_ASSISTANT)}角色'}, status=status.HTTP_500_INTERNAL_SERVER_ERROR)

        deleted_count, _ = UserRole.objects.filter(
            user=student,
            role=assistant_role,
            scope_type='class',
            scope_id=int(class_id),
        ).delete()

        if deleted_count == 0:
            return Response({'detail': f'该学生在此班级没有{get_role_display_name(ROLE_LEVEL_ASSISTANT)}身份'}, status=status.HTTP_404_NOT_FOUND)

        # 若撤销后该学生已无任何助理角色，清除 current_role（若当前是助理身份）
        if student.current_role_id == assistant_role.id:
            if not UserRole.objects.filter(user=student, role=assistant_role).exists():
                student.current_role = None
                student.save(update_fields=['current_role'])

        cls_name = ''
        try:
            cls_name = Class.objects.get(pk=class_id).name
        except Class.DoesNotExist:
            pass

        log_action(
            user=request.user,
            action='assistant_revoke',
            module=OperationLog.MODULE_USERS,
            level=OperationLog.LEVEL_WARNING,
            target_type='user',
            target_id=student.id,
            target_repr=student.username,
            extra={'class_id': class_id, 'class_name': cls_name},
            request=request,
        )

        return Response({'detail': f'已撤销{get_role_display_name(ROLE_LEVEL_ASSISTANT)}身份'})


class AssistantTaskListAPIView(ListAPIView):
    """
    GET /api/v1/review/assistant-tasks/
    学生助理（level==1）查看自己负责班级的待评分提交列表。
    评审老师及以上也可访问（查看助理任务概况）。
    """
    permission_classes = [IsAuthenticated]
    serializer_class = StudentSubmissionSerializer

    def get_queryset(self):
        user = self.request.user
        assigned_submission_ids = list(
            ReviewAssignment.objects.filter(
                reviewer=user, status__in=['assigned', 'completed']
            ).values_list('submission_id', flat=True).distinct()
        )
        if not assigned_submission_ids:
            return StudentSubmission.objects.none()

        return StudentSubmission.objects.filter(
            id__in=assigned_submission_ids,
        ).exclude(
            status='draft',
        ).exclude(
            via_late_channel=True, status='submitted'
        ).select_related('project', 'user').order_by('-updated_at')

class AssistantScoreAPIView(APIView):
    """
    POST /api/v1/review/assistant-tasks/<submission_id>/score/
    学生助理提交评分（需验证提交的学生属于助理管辖的班级）。
    复用现有 ScoreRecord，round_type 逻辑与普通评审一致。

    请求体::

        {
            "indicator_id": 1,
            "score": 85.5,
            "comment": "评语"
        }
    """
    permission_classes = [IsAuthenticated]

    def post(self, request, pk):
        # 仅 student_assistant（level==1）可使用此接口；level>=2 用普通评分接口
        if not user_level_at_least(request.user, 1):
            return Response({'detail': '无评分权限'}, status=status.HTTP_403_FORBIDDEN)
        if user_level_at_least(request.user, 2):
            return Response({'detail': f'{get_role_display_name(ROLE_LEVEL_COUNSELOR)}请使用标准评分接口（/review/submissions/<id>/scores/）'}, status=status.HTTP_400_BAD_REQUEST)

        try:
            sub = StudentSubmission.objects.select_related('project', 'user').get(pk=pk)
        except StudentSubmission.DoesNotExist:
            return Response({'detail': '提交不存在'}, status=status.HTTP_404_NOT_FOUND)

        if sub.user_id == request.user.id:
            return Response({'detail': '不可评审自己的提交'}, status=status.HTTP_403_FORBIDDEN)

        assignment = _get_active_assignment_for_submission(request.user, sub)
        if assignment is None:
            return Response({'detail': '您未被分配该提交的评审任务'}, status=status.HTTP_403_FORBIDDEN)
        if assignment.role_type != 'assistant':
            return Response({'detail': '当前任务类型不允许通过助理评分接口提交'}, status=status.HTTP_403_FORBIDDEN)

        if sub.status not in ('submitted', 'under_review'):
            return Response({'detail': '当前状态不可评分'}, status=status.HTTP_400_BAD_REQUEST)

        indicator_id = request.data.get('indicator_id')
        raw_score = request.data.get('score')
        comment = request.data.get('comment', '')

        if indicator_id is None or raw_score is None:
            return Response({'detail': '缺少 indicator_id 或 score'}, status=status.HTTP_400_BAD_REQUEST)

        try:
            ind = EvalIndicator.objects.get(project=sub.project_id, pk=indicator_id, score_source='self')
        except EvalIndicator.DoesNotExist:
            return Response({'detail': '指标不存在或不在题目化审核范围'}, status=status.HTTP_404_NOT_FOUND)
        score, score_error = _validate_indicator_score(ind, raw_score)
        if score_error:
            return Response({'detail': score_error}, status=status.HTTP_400_BAD_REQUEST)

        # round_type 由分配任务确定
        round_type = assignment.round_type
        rule = getattr(sub.project, 'review_rule', None)
        if _is_single_review(rule) and _single_mode_conflict_exists(
            submission=sub,
            indicator=ind,
            reviewer_id=request.user.id,
            round_type=round_type,
        ):
            return Response({'detail': '单评模式下同一指标仅允许1条常规评分记录'}, status=status.HTTP_400_BAD_REQUEST)

        ctx = _build_score_context(request.user)
        rec, created = ScoreRecord.objects.update_or_create(
            submission=sub, indicator=ind, reviewer=request.user, round_type=round_type,
            defaults={
                'score': score,
                'comment': comment,
                'score_channel': 'assignment',
                **ctx,
            },
        )
        recompute_submission_final_score(sub)
        _maybe_create_single_review_confirm(sub, request.user, assignment)
        _maybe_mark_assignment_completed(sub, request.user, assignment)
        _maybe_finalize_submission_status(sub)

        log_action(
            user=request.user,
            action='assistant_score_submit',
            module=OperationLog.MODULE_SCORING,
            level=OperationLog.LEVEL_NOTICE,
            target_type='score_record',
            target_id=rec.id,
            target_repr=f'提交#{sub.id} · 指标#{ind.id}',
            extra={
                'score': str(score),
                'round_type': round_type,
                'score_channel': 'assignment',
                'logical_round_label': _format_logical_round_label(round_type),
                'role': 'student_assistant',
                **ctx,
            },
            request=request,
        )

        record_payload = ScoreRecordSerializer(rec).data
        record_payload = _attach_logical_round_fields([record_payload])[0]
        return Response(record_payload, status=status.HTTP_201_CREATED if created else status.HTTP_200_OK)
